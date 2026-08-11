from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import PermissionDenied
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Sum
import datetime
from .models import Vendor, Product, PurchaseOrder, ProductLot, StockTransaction
from .serializers import VendorSerializer, ProductSerializer, PurchaseOrderSerializer, ProductLotSerializer, StockTransactionSerializer

class InventoryBaseViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        
        # If user is a superuser, or has the "Owner" role, allow passing center_id parameter
        is_owner = user.is_superuser or (user.role and user.role.name.lower() == 'owner')
        center_id = self.request.query_params.get('center_id')

        if is_owner:
            if center_id:
                return queryset.filter(center_id=center_id)
            return queryset
        else:
            # Regular staff only see items for their assigned centers
            if user.centers.exists():
                return queryset.filter(center__in=user.centers.all())
            elif user.center:
                return queryset.filter(center=user.center)
            return queryset.none()

    def perform_create(self, serializer):
        user = self.request.user
        is_owner = user.is_superuser or (user.role and user.role.name.lower() == 'owner')

        if not is_owner:
            # For non-owners, resolve center from request — honor which center they're working in
            center_id_param = (
                self.request.data.get('center') or
                self.request.data.get('center_id') or
                self.request.query_params.get('center_id')
            )
            if center_id_param:
                try:
                    center_id_param = int(center_id_param)
                except (TypeError, ValueError):
                    center_id_param = None

            if user.centers.exists():
                allowed_centers = user.centers.all()
                if center_id_param:
                    matched = allowed_centers.filter(id=center_id_param).first()
                    if not matched:
                        raise PermissionDenied("You are not assigned to this center.")
                    serializer.save(center=matched)
                else:
                    if allowed_centers.count() == 1:
                        serializer.save(center=allowed_centers.first())
                    else:
                        raise PermissionDenied("Center is required.")
            elif getattr(user, 'center', None):
                serializer.save(center=user.center)
            else:
                raise PermissionDenied("You are not assigned to any center.")
        else:
            serializer.save()


class VendorViewSet(InventoryBaseViewSet):
    queryset = Vendor.objects.order_by('id')
    serializer_class = VendorSerializer

    def destroy(self, request, *args, **kwargs):
        vendor = self.get_object()
        if vendor.purchase_orders.exists():
            return Response({"error": "Cannot delete vendor with existing purchase orders. Please disable or hide the vendor instead."}, status=400)
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=['get'])
    def bulk_upload_template(self, request):
        import io
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return Response({'error': 'openpyxl not installed.'}, status=400)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vendors Import"
        
        headers = [
            'vendorCode', 'vendorName', 'shortName', 'phone', 'email',
            'gstNumber', 'panNumber', 'address', 'city', 'state', 'pinCode'
        ]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
            
        help_text = [
            'Optional (for update)', 'Required', 'Optional', 'Optional', 'Optional',
            'Optional', 'Optional', 'Optional', 'Optional', 'Optional', 'Optional'
        ]
        ws.append(help_text)
        
        help_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
        help_font = Font(color="15803D", italic=True)
        for cell in ws[2]:
            cell.fill = help_fill
            cell.font = help_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        example_row = [
            'VND-001', 'Acme Supplies', 'Acme', '9876543210', 'contact@acme.com',
            '22AAAAA0000A1Z5', 'AAAAA0000A', '123 Market St', 'Mumbai', 'Maharashtra', '400001'
        ]
        ws.append(example_row)
            
        response_io = io.BytesIO()
        wb.save(response_io)
        response_io.seek(0)
        
        from django.http import HttpResponse
        response = HttpResponse(response_io.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="vendors_import_template.xlsx"'
        return response

    @action(detail=False, methods=['post'])
    def bulk_upload(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No file uploaded'}, status=400)
        try:
            import pandas as pd
            if file_obj.name.endswith('.csv'):
                df = pd.read_csv(file_obj)
            elif file_obj.name.endswith('.xlsx') or file_obj.name.endswith('.xls'):
                df = pd.read_excel(file_obj)
            else:
                return Response({'error': 'Unsupported file format.'}, status=400)

            df = df.fillna('')
            def normalize_col(c):
                return str(c).strip().lower().replace(' ', '').replace('_', '')
            df.columns = [normalize_col(c) for c in df.columns]
            records = df.to_dict('records')

            user = request.user
            from salon_admin.models import Center
            center_id_param = request.data.get('center_id') or request.query_params.get('center_id')
            if center_id_param and str(center_id_param).lower() != 'null':
                try:
                    all_centers = [Center.objects.get(id=int(center_id_param))]
                except Center.DoesNotExist:
                    all_centers = list(Center.objects.all())
            else:
                all_centers = list(Center.objects.all())
                
            if not all_centers:
                return Response({"error": "No centers available."}, status=400)

            created_count = 0
            updated_count = 0
            skipped_count = 0
            errors = []

            from django.db import transaction as db_transaction
            with db_transaction.atomic():
                for idx, row in enumerate(records, start=2):
                    name = str(row.get('vendorname', row.get('name', ''))).strip()
                    if not name or name in ('nan', 'None'):
                        errors.append(f'Row {idx}: Skipped – Vendor Name is empty.')
                        skipped_count += 1
                        continue

                    vendor_code = str(row.get('vendorcode', row.get('code', ''))).strip()
                    if vendor_code in ('nan', 'None', '0'): vendor_code = ''
                    short_name = str(row.get('shortname', '')).strip()
                    phone = str(row.get('phone', '')).strip()
                    if phone in ('nan', 'None'): phone = ''
                    email = str(row.get('email', '')).strip()
                    if email in ('nan', 'None'): email = ''
                    cst_number = str(row.get('gstnumber', row.get('gst', row.get('cstnumber', '')))).strip()
                    pan_number = str(row.get('pannumber', row.get('pan', ''))).strip()
                    address = str(row.get('address', '')).strip()
                    city = str(row.get('city', '')).strip()
                    state = str(row.get('state', '')).strip()
                    pin_code = str(row.get('pincode', row.get('pin', ''))).strip()
                    if pin_code in ('nan', 'None', '0'): pin_code = ''

                    for center in all_centers:
                        existing = None
                        if vendor_code:
                            existing = Vendor.objects.filter(vendor_code__iexact=vendor_code, center=center).first()
                        if not existing:
                            existing = Vendor.objects.filter(name__iexact=name, center=center).first()

                        if existing:
                            existing.name = name
                            if vendor_code: existing.vendor_code = vendor_code
                            if short_name: existing.short_name = short_name
                            if phone: existing.phone = phone
                            if email: existing.email = email
                            if cst_number: existing.cst_number = cst_number
                            if pan_number: existing.pan_number = pan_number
                            if address: existing.address = address
                            if city: existing.city = city
                            if state: existing.state = state
                            if pin_code: existing.pin_code = pin_code
                            existing.save()
                            updated_count += 1
                        else:
                            Vendor.objects.create(
                                center=center,
                                vendor_code=vendor_code,
                                name=name,
                                short_name=short_name,
                                phone=phone,
                                email=email,
                                cst_number=cst_number,
                                pan_number=pan_number,
                                address=address,
                                city=city,
                                state=state,
                                pin_code=pin_code
                            )
                            created_count += 1

            result = {'message': f'Upload complete: {created_count} created, {updated_count} updated, {skipped_count} skipped.', 
                      'created': created_count, 'updated': updated_count, 'skipped': skipped_count}
            if errors:
                result['warnings'] = errors[:20]
            return Response(result)

        except Exception as e:
            import traceback
            return Response({'error': str(e), 'detail': traceback.format_exc()}, status=400)

from django.utils.decorators import method_decorator
from django.utils.decorators import method_decorator
from django.views.decorators.vary import vary_on_headers

class ProductViewSet(InventoryBaseViewSet):
    queryset = Product.objects.all().select_related('center').prefetch_related('lots')
    serializer_class = ProductSerializer

    @method_decorator(vary_on_headers('Authorization'))
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    def perform_update(self, serializer):
        instance = serializer.save()
        update_all = self.request.query_params.get('update_all_centers', 'false').lower() == 'true'
        if update_all:
            match_kwargs = {}
            if instance.product_code:
                match_kwargs['product_code__iexact'] = instance.product_code
            elif instance.product_id_str:
                match_kwargs['product_id_str__iexact'] = instance.product_id_str
            else:
                match_kwargs['name__iexact'] = instance.name
            
            # Find matching products in other centers and update them
            from django.db import transaction
            with transaction.atomic():
                others = Product.objects.filter(**match_kwargs).exclude(id=instance.id)
                for other in others:
                    other.name = instance.name
                    other.brand = instance.brand
                    other.category = instance.category
                    other.sub_category = instance.sub_category
                    other.vendor_name = instance.vendor_name
                    other.price = instance.price
                    other.gst_percent = instance.gst_percent
                    other.barcode = instance.barcode
                    other.sac_code = instance.sac_code
                    other.reorder_level = instance.reorder_level
                    other.reorder_quantity = instance.reorder_quantity
                    other.is_active = instance.is_active
                    other.save()

    def destroy(self, request, *args, **kwargs):
        product = self.get_object()
        if product.transactions.exists() or product.purchaseorderitem_set.exists():
            return Response({"error": "Cannot delete product that has stock history or purchase orders. Please mark it as inactive instead."}, status=400)
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=['post'])
    def bulk_upload(self, request):
        """Bulk upload/update products from Excel or CSV file.
        Supported columns: productID, productName, category, subCategory,
        vendor, defaultPri, sac, barcode, productCo, brand, tax, active
        If productID matches an existing product it is updated; otherwise a new product is created.
        """
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No file uploaded'}, status=400)
        try:
            import pandas as pd
            if file_obj.name.endswith('.csv'):
                df = pd.read_csv(file_obj)
            elif file_obj.name.endswith('.xlsx') or file_obj.name.endswith('.xls'):
                df = pd.read_excel(file_obj)
            else:
                return Response({'error': 'Unsupported file format. Please upload a .csv or .xlsx file.'}, status=400)

            df = df.fillna('')
            # Normalize column names: strip whitespace, lowercase, remove spaces and underscores
            def normalize_col(c):
                return str(c).strip().lower().replace(' ', '').replace('_', '')
            df.columns = [normalize_col(c) for c in df.columns]
            records = df.to_dict('records')

            user = request.user
            is_owner = user.is_superuser or (user.role and user.role.name.lower() == 'owner')

            from salon_admin.models import Center
            center_id_param = request.data.get('center_id') or request.query_params.get('center_id')
            if center_id_param and str(center_id_param).lower() != 'null':
                try:
                    all_centers = [Center.objects.get(id=int(center_id_param))]
                except Center.DoesNotExist:
                    all_centers = list(Center.objects.all())
            else:
                # Evaluate immediately to a list — avoids lazy-queryset reads inside atomic() causing deadlocks
                all_centers = list(Center.objects.all())
                
            if not all_centers:
                return Response({"error": "No centers available."}, status=400)

            # Helper to safely parse numeric values
            def safe_float(val, default=0.0):
                try:
                    v = str(val).strip()
                    return float(v) if v not in ('', 'nan', 'None') else default
                except (ValueError, TypeError):
                    return default

            def safe_int(val, default=0):
                try:
                    v = str(val).strip()
                    return int(float(v)) if v not in ('', 'nan', 'None') else default
                except (ValueError, TypeError):
                    return default

            # Map category: anything that is not 'Professional' defaults to 'Retail'
            VALID_CATEGORIES = {'retail': 'Retail', 'professional': 'Professional'}

            created_count = 0
            updated_count = 0
            skipped_count = 0
            errors = []

            from django.db import transaction as db_transaction
            with db_transaction.atomic():
                for idx, row in enumerate(records, start=2):  # start=2 for Excel row numbers
                    # --- productName (required) ---
                    name = str(row.get('productname', row.get('name', ''))).strip()
                    if not name or name in ('nan', 'None'):
                        errors.append(f'Row {idx}: Skipped – productName is empty.')
                        skipped_count += 1
                        continue

                    # --- productID: check if existing product ---
                    product_id_raw = row.get('productid', row.get('id', ''))
                    product_id = safe_int(product_id_raw, default=0)
                    product_id_str_val = str(product_id) if product_id else ''

                    # --- Read all fields from template ---
                    category = str(row.get('category', '')).strip()
                    sub_category = str(row.get('subcategory', '')).strip()
                    vendor_name = str(row.get('vendor', row.get('vendorname', ''))).strip()
                    is_active = str(row.get('active', row.get('isactive', 'True'))).strip().lower() in ['true', '1', 'yes', 't', 'y']

                    price = safe_float(row.get('defaultprice', row.get('defaultpri', row.get('price', 0))))
                    sac_code = str(row.get('sac', row.get('saccode', ''))).strip()
                    if sac_code in ('', 'nan', 'None', '0'):
                        sac_code = ''

                    barcode = str(row.get('barcode', '')).strip()
                    if barcode in ('', 'nan', 'None', '0'):
                        barcode = ''

                    product_code = str(row.get('productcode', row.get('productco', ''))).strip()
                    if product_code in ('', 'nan', 'None', '0'):
                        product_code = ''

                    brand = str(row.get('brand', '')).strip()
                    if brand in ('nan', 'None'):
                        brand = ''

                    gst_percent = safe_float(row.get('tax', row.get('gstpercent', 0.18)))
                    # tax column often stored as 0.18 (fraction) – convert to percent if <= 1
                    if 0 < gst_percent <= 1:
                        gst_percent = gst_percent * 100

                    reorder_level = safe_int(row.get('reorderlevel', 0))
                    reorder_quantity = safe_int(row.get('reorderquantity', 0))
                    current_stock = safe_int(row.get('currentstock', row.get('stock', 0)))

                    # Apply to ALL centers

                    for center in all_centers:
                        existing = None

                        # Match priority: product_code > product_id_str > name (all scoped to this center)
                        if product_code:
                            existing = Product.objects.filter(product_code__iexact=product_code, center=center).first()
                        if not existing and product_id_str_val:
                            existing = Product.objects.filter(product_id_str__iexact=product_id_str_val, center=center).first()
                        if not existing and name:
                            existing = Product.objects.filter(name__iexact=name, center=center).first()

                        if existing:
                            # Update existing product (preserve stock)
                            existing.name = name
                            if product_id_str_val:
                                existing.product_id_str = product_id_str_val
                            if product_code:
                                existing.product_code = product_code
                            if category:
                                existing.category = category
                            if sub_category:
                                existing.sub_category = sub_category
                            if vendor_name:
                                existing.vendor_name = vendor_name
                            if brand:
                                existing.brand = brand
                            if price:
                                existing.price = price
                            if sac_code:
                                existing.sac_code = sac_code
                            if barcode:
                                existing.barcode = barcode
                            existing.gst_percent = gst_percent
                            existing.is_active = is_active
                            if reorder_level:
                                existing.reorder_level = reorder_level
                            if reorder_quantity:
                                existing.reorder_quantity = reorder_quantity
                            existing.save()
                            updated_count += 1
                        else:
                            # Create new product for this center
                            Product.objects.create(
                                center=center,
                                product_id_str=product_id_str_val,
                                product_code=product_code,
                                name=name,
                                category=category,
                                sub_category=sub_category,
                                vendor_name=vendor_name,
                                brand=brand,
                                price=price if price else None,
                                sac_code=sac_code,
                                barcode=barcode,
                                gst_percent=gst_percent,
                                is_active=is_active,
                                reorder_level=reorder_level,
                                reorder_quantity=reorder_quantity,
                                current_stock=current_stock,
                            )
                            created_count += 1

            msg = f'Upload complete: {created_count} created, {updated_count} updated, {skipped_count} skipped.'
            result = {'message': msg, 'created': created_count, 'updated': updated_count, 'skipped': skipped_count}
            if errors:
                result['warnings'] = errors[:20]  # Return first 20 warnings
            return Response(result)

        except Exception as e:
            import traceback
            return Response({'error': str(e), 'detail': traceback.format_exc()}, status=400)

    @action(detail=False, methods=['get'])
    def bulk_upload_template(self, request):
        import io
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return Response({'error': 'openpyxl not installed.'}, status=400)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products Import"
        
        headers = [
            'productID', 'productName', 'category', 'subCategory', 'vendor',
            'defaultPrice', 'sac', 'barcode', 'productCode', 'brand', 'tax', 'active',
            'currentStock', 'reorderLevel', 'reorderQuantity'
        ]
        
        ws.append(headers)
        
        # Styling
        header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
            
        # Helper row
        help_text = [
            'Leave blank for new', 'Required', 'Retail / Professional', 'Optional', 'Optional',
            'Number (e.g. 100.00)', 'Optional', 'Optional', 'Optional', 'Optional', 'Fraction (e.g. 0.18)', 'TRUE / FALSE',
            'Number (e.g. 10)', 'Number', 'Number'
        ]
        ws.append(help_text)
        
        help_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
        help_font = Font(color="15803D", italic=True)
        for cell in ws[2]:
            cell.fill = help_fill
            cell.font = help_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # Example row
        example_row = [
            '', 'Loreal Shampoo', 'Retail', 'Hair Care', 'Loreal',
            1500.00, '9902', '123456789', 'PRD-001', 'Loreal', 0.18, 'TRUE',
            50, 10, 20
        ]
        ws.append(example_row)
            
        response_io = io.BytesIO()
        wb.save(response_io)
        response_io.seek(0)
        
        from django.http import HttpResponse
        response = HttpResponse(response_io.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="products_import_template.xlsx"'
        return response

    @action(detail=False, methods=['post'])
    def checkout(self, request):
        from django.db import transaction
        items = request.data.get('items', [])
        user = request.user
        # Find user center
        center = user.center if user.center else (user.centers.first() if user.centers.exists() else None)
        if not center and not user.is_superuser:
            return Response({"error": "No center assigned to user"}, status=400)
            
        # if superuser, get center from request data or use first center
        if user.is_superuser or (user.role and user.role.name == 'Owner'):
            center_id = request.data.get('center_id')
            if center_id:
                from salon_admin.models import Center
                center = Center.objects.get(id=center_id)
        
        created_transactions = []
        with transaction.atomic():
            for item in items:
                # Lock row to prevent race conditions during concurrent checkouts
                product = Product.objects.select_for_update().get(id=item['product_id'])
                
                if product.center != center:
                    return Response({"error": f"Product {product.name} does not belong to the selected center."}, status=400)
                    
                qty = Decimal(str(item['quantity']))
                
                if qty > product.current_stock:
                    return Response({"error": f"Cannot checkout {qty} of {product.name}. Only {product.current_stock} in stock!"}, status=400)
                
                product.current_stock -= qty
                product.save(update_fields=['current_stock'])
                
                tx = StockTransaction.objects.create(
                    product=product,
                    center=center,
                    transaction_type='CHECKOUT',
                    quantity_change=-qty,
                    created_by=user
                )
                created_transactions.append(tx)
                
        return Response({"status": "success", "message": f"Checked out {len(created_transactions)} items"})

    @action(detail=False, methods=['post'])
    def audit(self, request):
        from django.db import transaction
        items = request.data.get('items', [])
        user = request.user
        center = user.center if user.center else (user.centers.first() if user.centers.exists() else None)
        if user.is_superuser or (user.role and user.role.name == 'Owner'):
            center_id = request.data.get('center_id')
            if center_id:
                from salon_admin.models import Center
                center = Center.objects.get(id=center_id)
        
        audits = []
        with transaction.atomic():
            for item in items:
                # Lock row to prevent race conditions during concurrent audits/checkouts
                product = Product.objects.select_for_update().get(id=item['product_id'])
                
                if product.center != center:
                    return Response({"error": f"Product {product.name} does not belong to the selected center."}, status=400)
                    
                physical_qty = Decimal(str(item['quantity']))
                diff = physical_qty - product.current_stock
                
                if diff != 0:
                    product.current_stock = physical_qty
                    product.save(update_fields=['current_stock'])
                    
                    tx = StockTransaction.objects.create(
                        product=product,
                        center=center,
                        transaction_type='AUDIT',
                        quantity_change=diff,
                        created_by=user
                    )
                    audits.append(tx)
                
        return Response({"status": "success", "message": f"Audited {len(audits)} items with changes"})

    @action(detail=False, methods=['get'])
    def stock_history(self, request):
        date_str = request.query_params.get('date') # Format: YYYY-MM-DD
        center_id = request.query_params.get('center_id')
        
        if not date_str:
            return Response({"error": "date parameter required"}, status=400)
            
        try:
            target_date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return Response({"error": "Invalid date format, use YYYY-MM-DD"}, status=400)
            
        products = self.get_queryset()
        if center_id:
            products = products.filter(center_id=center_id)
            
        # Target date end of day
        naive_dt = datetime.datetime.combine(target_date, datetime.time.max)
        target_datetime = naive_dt
        
        # Optimize: get future transaction sums for all products in one query
        from django.db.models import Sum, Q
        
        # Annotate each product with the sum of its future transactions
        products = products.annotate(
            future_sum=Sum(
                'transactions__quantity_change',
                filter=Q(transactions__created_at__gt=target_datetime)
            )
        )
        
        results = []
        for p in products:
            # We want stock as of target_datetime.
            # Current stock = stock as of NOW
            # Stock as of target = Current stock - sum(transactions after target_datetime)
            future_sum = p.future_sum or 0
            historical_stock = p.current_stock - future_sum
            
            results.append({
                "product_id": p.id,
                "name": p.name,
                "brand": p.brand,
                "category": p.category,
                "price": p.price,
                "historical_stock": historical_stock
            })
            
        return Response(results)

class ProductLotViewSet(InventoryBaseViewSet):
    queryset = ProductLot.objects.order_by('-id')
    serializer_class = ProductLotSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        product_id = self.request.data.get('product')
        if not product_id:
            raise PermissionDenied("Product ID is required")
        product = Product.objects.get(id=product_id)
        
        user = self.request.user
        is_owner = user.is_superuser or (user.role and user.role.name == 'Owner')
        if not is_owner:
            if user.centers.exists() and product.center not in user.centers.all():
                raise PermissionDenied("You do not have access to the product's center.")
            elif user.center and product.center != user.center:
                raise PermissionDenied("You do not have access to the product's center.")
                
        serializer.save(product=product)

class PurchaseOrderViewSet(InventoryBaseViewSet):
    queryset = PurchaseOrder.objects.all().select_related('vendor', 'center').prefetch_related('items__product').order_by('-created_at')
    serializer_class = PurchaseOrderSerializer

    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        po = self.get_object()
        if po.status == 'Delivered':
            return Response({"error": "Cannot delete a delivered purchase order. Please refund or create a manual adjustment instead."}, status=400)
        return super().destroy(request, *args, **kwargs)

class StockTransactionViewSet(InventoryBaseViewSet):
    queryset = StockTransaction.objects.all().select_related('product', 'product__center', 'center').order_by('-created_at')
    serializer_class = StockTransactionSerializer

from rest_framework.views import APIView
from django.db.models import F

class LowStockAlertView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        center_id = request.query_params.get('center_id')
        qs = Product.objects.filter(current_stock__lte=F('reorder_level')).select_related('center')
        
        user = request.user
        is_owner = user.is_superuser or (user.role and user.role.name == 'Owner')
        if not is_owner:
            if user.centers.exists():
                qs = qs.filter(center__in=user.centers.all())
            elif user.center:
                qs = qs.filter(center=user.center)
            else:
                qs = qs.none()
                
        if center_id:
            qs = qs.filter(center_id=center_id)
            
        data = []
        for p in qs:
            data.append({
                'id': p.id,
                'name': p.name,
                'stock': p.current_stock,
                'reorder_level': p.reorder_level,
                'center': p.center.display_name or p.center.center_name if p.center else None
            })
        return Response(data)


