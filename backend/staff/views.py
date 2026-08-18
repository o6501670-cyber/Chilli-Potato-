import datetime
from decimal import Decimal

import openpyxl
from django.contrib.auth.hashers import check_password
from django.db.models import Count, Q, Sum
from rest_framework import permissions, status, viewsets
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response

from salon_admin.models import Center

from .models import (
    Designation,
    PayrollRecord,
    ServiceLog,
    StaffConsumptionLog,
    StaffMember,
    StaffToolTracker,
    StaffTransfer,
)
from .serializers import (
    DesignationSerializer,
    PayrollRecordSerializer,
    ServiceLogSerializer,
    StaffConsumptionLogSerializer,
    StaffMemberSerializer,
    StaffToolTrackerSerializer,
    StaffTransferSerializer,
)
from .utils import sync_staff_transfers_and_tools


class DesignationViewSet(viewsets.ModelViewSet):
    queryset = Designation.objects.all().order_by('name')
    serializer_class = DesignationSerializer
    permission_classes = [permissions.IsAuthenticated]

class StaffMemberViewSet(viewsets.ModelViewSet):
    serializer_class = StaffMemberSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        from datetime import date

        from django.db.models import Exists, OuterRef

        from .models import StaffToolTracker
        
        # Optimize N+1 queries by selecting related center and annotating overdue tools
        overdue_tools = StaffToolTracker.objects.filter(
            staff=OuterRef('pk'),
            status='Taken',
            expected_return_date__lt=date.today()
        )
        queryset = StaffMember.objects.all().select_related('center').annotate(
            has_overdue_tools_annotated=Exists(overdue_tools)
        ).order_by('first_name')
        role = getattr(user, 'role', None)
        is_owner = IsOwner.check_is_owner(user)
        perms = getattr(role, 'permissions', {}) or {}

        if not is_owner and not perms.get('all_centers', False):
            if user.centers.exists():
                queryset = queryset.filter(center__in=user.centers.all())
            elif hasattr(user, 'center') and user.center:
                queryset = queryset.filter(center=user.center)

        center_id = self.request.query_params.get('center_id')
        if center_id:
            queryset = queryset.filter(center_id=center_id)

        include_inactive = self.request.query_params.get('include_inactive', 'false')
        if self.action == 'list' and include_inactive != 'true':
            queryset = queryset.filter(is_active=True)

        return queryset

    @action(detail=False, methods=['post'])
    def sync_transfers(self, request):
        """Manually trigger transfer/tool expiry sync. Call from a scheduled task or management command."""
        from .utils import sync_staff_transfers_and_tools
        try:
            sync_staff_transfers_and_tools()
            return Response({'status': 'sync complete'})
        except Exception as e:
            return Response({'error': str(e)}, status=500)


    @action(detail=False, methods=['get'])
    def bulk_upload_template(self, request):
        import io
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            return Response({'error': 'openpyxl not installed.'}, status=400)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Staff Import"
        
        headers = [
            'staffCode', 'firstName', 'lastName', 'gender', 'designation', 'center',
            'salary', 'joiningDate', 'phone', 'email', 'address', 'city', 'state', 'pinCode'
        ]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
            
        help_text = [
            'Optional (for update)', 'Required', 'Optional', 'Male/Female', 'Optional', 'Required',
            'Optional', 'YYYY-MM-DD', 'Optional', 'Optional', 'Optional', 'Optional', 'Optional', 'Optional'
        ]
        ws.append(help_text)
        
        help_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
        help_font = Font(color="15803D", italic=True)
        for cell in ws[2]:
            cell.fill = help_fill
            cell.font = help_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        example_row = [
            'STF-001', 'John', 'Doe', 'Male', 'Senior Stylist', 'Main Center',
            '25000', '2023-01-15', '9876543210', 'john@example.com', '123 Main St', 'Mumbai', 'Maharashtra', '400001'
        ]
        ws.append(example_row)
            
        response_io = io.BytesIO()
        wb.save(response_io)
        response_io.seek(0)
        
        from django.http import HttpResponse
        response = HttpResponse(response_io.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="staff_import_template.xlsx"'
        return response

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def bulk_upload(self, request):
        if 'file' not in request.FILES:
            return Response({'detail': 'No file provided'}, status=400)
            
        excel_file = request.FILES['file']
        
        filename = excel_file.name.lower()
        rows = []
        if filename.endswith('.csv'):
            import csv
            from io import StringIO
            try:
                decoded_file = excel_file.read().decode('utf-8')
                io_string = StringIO(decoded_file)
                reader = csv.reader(io_string)
                rows = list(reader)
            except Exception as e:
                return Response({'detail': f'Error reading CSV file: {e!s}'}, status=400)
        else:
            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                sheet = wb.active
                rows = list(sheet.iter_rows(values_only=True))
            except Exception as e:
                return Response({'detail': f'Error reading Excel file: {e!s}'}, status=400)

        header_idx = 0
        for idx, row in enumerate(rows):
            if any(str(cell).strip() for cell in row if cell is not None):
                header_idx = idx
                break
                
        if len(rows) <= header_idx + 1:
            return Response({'detail': 'File is empty or missing headers'}, status=400)
            
        header = [str(h).strip().lower().replace(' ', '_') if h else '' for h in rows[header_idx]]
        
        user = request.user
        center_id_param = request.data.get('center_id') or request.query_params.get('center_id')
        if center_id_param and str(center_id_param).lower() != 'null':
            try:
                all_centers = [Center.objects.get(id=int(center_id_param))]
            except Center.DoesNotExist:
                all_centers = list(Center.objects.all())
        else:
            all_centers = list(Center.objects.all())
            
        if not all_centers:
            new_center = Center.objects.create(center_name="Main Center")
            all_centers = [new_center]

        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []
        
        from django.db import transaction as db_transaction
        with db_transaction.atomic():
            for i, row in enumerate(rows[header_idx+1:], start=header_idx+2):
                if not any(row):
                    continue
                
                row_data = dict(zip(header, row))
                
                first_name = str(row_data.get('firstname', row_data.get('first_name', ''))).strip()
                last_name = str(row_data.get('lastname', row_data.get('last_name', ''))).strip()
                
                if not first_name:
                    raw_name = str(row_data.get('name') or row_data.get('full_name') or row_data.get('staff_name') or '').strip()
                    if raw_name:
                        parts = raw_name.split(' ', 1)
                        first_name = parts[0]
                        last_name = parts[1] if len(parts) > 1 else ''
                
                if not first_name or first_name in ('nan', 'None'):
                    errors.append(f"Row {i}: Skipped – First Name is empty.")
                    skipped_count += 1
                    continue
                    
                staff_code = str(row_data.get('staffcode', row_data.get('staff_code', ''))).strip()
                if staff_code in ('nan', 'None', '0'): staff_code = ''
                
                phone = str(row_data.get('phone', row_data.get('phone_number', row_data.get('mobile', '')))).strip()
                phone = phone.removesuffix('.0')
                if phone in ('nan', 'None'): phone = ''
                
                email = str(row_data.get('email', '')).strip()
                if email in ('nan', 'None'): email = ''
                
                address = str(row_data.get('address', '')).strip()
                city = str(row_data.get('city', '')).strip()
                state = str(row_data.get('state', '')).strip()
                pin_code = str(row_data.get('pincode', row_data.get('pin', ''))).strip()
                if pin_code in ('nan', 'None', '0'): pin_code = ''
                
                loc_name = str(row_data.get('center') or row_data.get('location_name') or row_data.get('location') or '').strip()
                center = None
                if loc_name and loc_name not in ('nan', 'None'):
                    center = Center.objects.filter(center_name__icontains=loc_name).first()
                if not center:
                    return Response({'error': f"Row {idx+1}: Center '{loc_name}' not found or empty."}, status=status.HTTP_400_BAD_REQUEST)
                    
                designation = str(row_data.get('designation', '')).strip()
                if designation in ('nan', 'None'): designation = ''
                
                gender = str(row_data.get('gender', '')).strip().capitalize()
                if gender not in ['Male', 'Female', 'Other']:
                    gender = 'Female'
                    
                raw_salary = row_data.get('salary', row_data.get('monthly_gross', 0))
                try:
                    salary = Decimal(str(raw_salary))
                except Exception:
                    salary = Decimal('0.0')
                    
                comm_perc = row_data.get('commission_percentage', row_data.get('comm._%', row_data.get('comm_%', 0)))
                try:
                    comm_perc = Decimal(str(comm_perc))
                except Exception:
                    comm_perc = Decimal('0.0')

                prod_comm = row_data.get('product_commission_percentage', row_data.get('prod._comm._%', row_data.get('prod_comm_%', 0)))
                try:
                    prod_comm = Decimal(str(prod_comm))
                except Exception:
                    prod_comm = Decimal('0.0')
                    
                raw_date = row_data.get('joindate', row_data.get('joiningdate', row_data.get('joining_date', '')))
                joining_date = None
                if raw_date and str(raw_date) not in ('nan', 'None', ''):
                    if hasattr(raw_date, 'date'):
                        joining_date = raw_date.date()
                    else:
                        try:
                            import datetime
                            joining_date = datetime.datetime.strptime(str(raw_date).strip(), '%d %b %Y').date()
                        except ValueError:
                            try:
                                joining_date = datetime.datetime.strptime(str(raw_date).strip(), '%Y-%m-%d').date()
                            except ValueError:
                                pass
                                
                existing = None
                if staff_code:
                    existing = StaffMember.objects.filter(staff_code__iexact=staff_code, center=center).first()
                if not existing and phone:
                    existing = StaffMember.objects.filter(phone=phone, center=center).first()
                if not existing:
                    existing = StaffMember.objects.filter(first_name__iexact=first_name, last_name__iexact=last_name, center=center).first()

                try:
                    if existing:
                        existing.first_name = first_name
                        if last_name and last_name not in ('nan', 'None'): existing.last_name = last_name
                        if staff_code: existing.staff_code = staff_code
                        if phone: existing.phone = phone
                        if email: existing.email = email
                        if address and address not in ('nan', 'None'): existing.address = address
                        if city and city not in ('nan', 'None'): existing.city = city
                        if state and state not in ('nan', 'None'): existing.state = state
                        if pin_code: existing.pin_code = pin_code
                        if designation: existing.designation = designation
                        existing.gender = gender
                        if salary: existing.salary = salary
                        if comm_perc: existing.commission_percentage = comm_perc
                        if prod_comm: existing.product_commission_percentage = prod_comm
                        if joining_date: existing.joining_date = joining_date
                        existing.save()
                        updated_count += 1
                    else:
                        StaffMember.objects.create(
                            center=center,
                            first_name=first_name,
                            last_name=last_name,
                            staff_code=staff_code,
                            phone=phone,
                            email=email,
                            address=address,
                            city=city,
                            state=state,
                            pin_code=pin_code,
                            designation=designation,
                            gender=gender,
                            salary=salary,
                            commission_percentage=comm_perc,
                            product_commission_percentage=prod_comm,
                            joining_date=joining_date,
                            is_active=True
                        )
                        created_count += 1
                except Exception as e:
                    errors.append(f"Row {i}: Error saving '{first_name}' - {e!s}")
                    skipped_count += 1
                    
        result = {
            'message': f'Upload complete: {created_count} created, {updated_count} updated, {skipped_count} skipped.',
            'created': created_count, 'updated': updated_count, 'skipped': skipped_count
        }
        if errors:
            result['warnings'] = errors[:20]
        return Response(result)

    def perform_create(self, serializer):
        user = self.request.user
        perms = getattr(user.role, 'permissions', {}) or {}
        is_owner = IsOwner.check_is_owner(user)
        
        if not is_owner and not perms.get('all_centers', False):
            center = serializer.validated_data.get('center')
            if center:
                if user.centers.exists() and center not in user.centers.all() or not user.centers.exists() and hasattr(user, 'center') and center != user.center:
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("You cannot create staff for this center.")
        serializer.save()

    @action(detail=False, methods=['get'])
    def activity_feed(self, request):
        feed = []
        
        center_id = request.query_params.get('center_id')
        
        services_qs = ServiceLog.objects.exclude(
            invoice__status__in=['cancelled', 'refunded']
        ).select_related('staff', 'invoice__client').order_by('-date', '-time')
        consumptions_qs = StaffConsumptionLog.objects.select_related('staff').order_by('-date', '-time')
        tools_qs = StaffToolTracker.objects.select_related('staff').order_by('-created_at')
        transfers_qs = StaffTransfer.objects.select_related('staff', 'to_center', 'from_center').order_by('-created_at')

        if center_id:
            services_qs = services_qs.filter(staff__center_id=center_id)
            consumptions_qs = consumptions_qs.filter(staff__center_id=center_id)
            tools_qs = tools_qs.filter(staff__center_id=center_id)
            transfers_qs = transfers_qs.filter(staff__center_id=center_id)
        
        # Limit each source to 20 to reduce data load — final sort will pick top 50 overall
        services = services_qs[:20]
        for s in services:
            dt = datetime.datetime.combine(s.date, s.time if s.time else datetime.time())
            client_display = s.client_name
            if s.invoice and s.invoice.client:
                client_display = f"{s.invoice.client.first_name} {s.invoice.client.last_name or ''}".strip()

            feed.append({
                'id': f"srv_{s.id}",
                'type': 'service',
                'staff_name': f"{s.staff.first_name} {s.staff.last_name or ''}".strip(),
                'title': 'Logged a Service',
                'description': '',
                'details': {
                    'Client': client_display,
                    'Item': s.service_name,
                    'Type': s.service_type,
                    'Price': f"₹{s.price}"
                },
                'timestamp': dt.isoformat()
            })
            
        consumptions = consumptions_qs[:20]
        for c in consumptions:
            dt = datetime.datetime.combine(c.date, c.time if c.time else datetime.time())
            feed.append({
                'id': f"cons_{c.id}",
                'type': 'consumption',
                'staff_name': f"{c.staff.first_name} {c.staff.last_name or ''}".strip(),
                'title': 'Consumed a Perk',
                'description': '',
                'details': {
                    'Item': c.service_name,
                    'Payment': c.payment_method,
                    'Cost': f"₹{c.amount}"
                },
                'timestamp': dt.isoformat()
            })
            
        tools = tools_qs[:20]
        for t in tools:
            feed.append({
                'id': f"tool_{t.id}",
                'type': 'tool',
                'staff_name': f"{t.staff.first_name} {t.staff.last_name or ''}".strip(),
                'title': 'Tool Assigned' if t.status == 'Taken' else 'Tool Returned',
                'description': '',
                'details': {
                    'Tool': t.tool_name,
                    'Qty': t.amount,
                    'Status': t.status,
                    'Return By': t.expected_return_date.strftime('%d %b %Y') if t.expected_return_date else 'N/A'
                },
                'timestamp': t.created_at.isoformat()
            })
            
        transfers = transfers_qs[:20]
        for tr in transfers:
            feed.append({
                'id': f"trans_{tr.id}",
                'type': 'transfer',
                'staff_name': f"{tr.staff.first_name} {tr.staff.last_name or ''}".strip(),
                'title': 'Staff Transfer',
                'description': '',
                'details': {
                    'Type': tr.transfer_type,
                    'From': (tr.from_center.display_name or tr.from_center.center_name) if tr.from_center else 'Unknown',
                    'To': (tr.to_center.display_name or tr.to_center.center_name) if tr.to_center else 'Unknown',
                    'Duration': f"{tr.start_date.strftime('%d %b %Y')} to {tr.end_date.strftime('%d %b %Y') if tr.end_date else 'Permanent'}"
                },
                'timestamp': tr.created_at.isoformat()
            })
            
        feed.sort(key=lambda x: x['timestamp'], reverse=True)
        return Response(feed[:50])

    @action(detail=False, methods=['get'])
    def commission_report(self, request):
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        if not start_date or not end_date:
            return Response({'error': 'start_date and end_date are required'}, status=400)

        center_id = request.query_params.get('center_id')
        staff_qs = self.get_queryset()
        if center_id:
            staff_qs = staff_qs.filter(center_id=center_id)

        logs = (
            ServiceLog.objects
            .filter(
                staff__in=staff_qs,
                date__gte=start_date,
                date__lte=end_date,
                invoice__status__in=['paid', 'partial'],
            )
            .values('staff_id', 'staff__first_name', 'staff__last_name',
                    'staff__commission_percentage', 'staff__product_commission_percentage',
                    'staff__salary', 'staff__center__display_name')
            .annotate(
                service_revenue=Sum('price', filter=Q(service_type__in=['Service', 'Membership', 'Package'])),
                product_revenue=Sum('price', filter=Q(service_type='Product')),
            )
        )

        report = []
        for row in logs:
            svc_rev = Decimal(str(row['service_revenue'] or 0))
            prod_rev = Decimal(str(row['product_revenue'] or 0))
            svc_comm = svc_rev * Decimal(str(row['staff__commission_percentage'] or 0)) / Decimal(100)
            prod_comm = prod_rev * Decimal(str(row['staff__product_commission_percentage'] or 0)) / Decimal(100)
            name = row['staff__first_name']
            if row['staff__last_name']:
                name += f" {row['staff__last_name']}"
            report.append({
                'staff_id': row['staff_id'],
                'staff_name': name,
                'center_name': row['staff__center__display_name'] or 'N/A',
                'salary': Decimal(str(row['staff__salary'] or 0)),
                'service_revenue': svc_rev,
                'product_revenue': prod_rev,
                'total_revenue': svc_rev + prod_rev,
                'service_commission': round(svc_comm, 2),
                'product_commission': round(prod_comm, 2),
                'total_commission': round(svc_comm + prod_comm, 2),
            })
        return Response(report)

    @action(detail=False, methods=['get'])
    def incentive_report(self, request):
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        if not start_date or not end_date:
            return Response({'error': 'start_date and end_date are required'}, status=400)

        staff_qs = self.get_queryset().select_related('center')

        logs = (
            ServiceLog.objects
            .filter(
                staff__in=staff_qs,
                date__gte=start_date,
                date__lte=end_date,
                invoice__status__in=['paid', 'partial']
            )
            .values(
                'staff_id',
                'staff__first_name',
                'staff__last_name',
                'staff__salary',
                'staff__center__display_name',
            )
            .annotate(total_sales=Sum('price'))
        )

        sales_by_staff = {row['staff_id']: row for row in logs}

        report = []
        
        # Check for dynamic overrides from the frontend
        override_multiplier = request.query_params.get('target_multiplier')
        override_percent = request.query_params.get('incentive_percent')
        
        if override_multiplier and override_percent:
            try:
                tiers = [(Decimal(str(override_multiplier)), Decimal(str(override_percent)))]
            except (ValueError, TypeError, Exception):
                tiers = []
        else:
            from finance.models import IncentiveConfig
            center_id_filter = request.query_params.get('center_id')
            incentive_configs = IncentiveConfig.objects.filter(
                category='multiplier_tier',
            ).order_by('-custom_percent')
            if center_id_filter:
                center_configs = incentive_configs.filter(center_id=center_id_filter)
                if center_configs.exists():
                    incentive_configs = center_configs
                else:
                    incentive_configs = incentive_configs.filter(center__isnull=True)

            tiers = [(Decimal(str(cfg.name)), Decimal(str(cfg.custom_percent))) for cfg in incentive_configs if _is_decimal(cfg.name)]
            if not tiers:
                tiers = [(Decimal('7.0'), Decimal('5.0')), (Decimal('6.0'), Decimal('4.0')), (Decimal('5.0'), Decimal('3.0'))]
            tiers.sort(key=lambda t: t[0], reverse=True)

        for member in staff_qs:
            row = sales_by_staff.get(member.id, {})
            total_sales = Decimal(str(row.get('total_sales') or 0))
            salary = Decimal(str(member.salary or 0))
            multiplier = total_sales / salary if salary > 0 else Decimal(0)

            incentive_percentage = Decimal(0)
            for threshold, pct in tiers:
                if multiplier >= threshold:
                    incentive_percentage = pct
                    break

            incentive_amount = total_sales * (incentive_percentage / Decimal('100.0'))

            report.append({
                'staff_id': member.id,
                'staff_name': str(member),
                'center_name': member.center.display_name if member.center else 'N/A',
                'salary': salary,
                'total_sales': round(total_sales, 2),
                'multiplier': round(multiplier, 2),
                'incentive_percentage': incentive_percentage,
                'incentive_amount': round(incentive_amount, 2),
            })

        return Response(report)

def _is_decimal(s):
    try:
        Decimal(str(s))
        return True
    except (TypeError, ValueError, Exception):
        return False


class ServiceLogViewSet(viewsets.ModelViewSet):
    serializer_class = ServiceLogSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        queryset = (
            ServiceLog.objects
            .exclude(invoice__status__in=['cancelled', 'refunded'])
            .select_related('staff', 'center', 'invoice', 'invoice__client')
            .order_by('-date', '-time')
        )
        role = getattr(user, 'role', None)
        is_owner = IsOwner.check_is_owner(user)
        perms = getattr(role, 'permissions', {}) or {}

        if not is_owner and not perms.get('all_centers', False):
            if user.centers.exists():
                queryset = queryset.filter(center__in=user.centers.all())
            elif hasattr(user, 'center') and user.center:
                queryset = queryset.filter(center=user.center)

        center_id = self.request.query_params.get('center_id')
        if center_id:
            queryset = queryset.filter(center_id=center_id)

        staff_id = self.request.query_params.get('staff_id')
        if staff_id:
            queryset = queryset.filter(staff_id=staff_id)

        client_name = self.request.query_params.get('client_name')
        if client_name:
            queryset = queryset.filter(client_name__icontains=client_name)

        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date and end_date:
            queryset = queryset.filter(date__range=[start_date, end_date])
        elif start_date:
            queryset = queryset.filter(date__gte=start_date)
        elif end_date:
            queryset = queryset.filter(date__lte=end_date)
        else:
            # Default: return last 30 days to prevent unbounded full-table scan on list
            from datetime import date, timedelta
            queryset = queryset.filter(date__gte=date.today() - timedelta(days=30))

        return queryset


    def perform_create(self, serializer):
        user = self.request.user
        perms = getattr(user.role, 'permissions', {}) or {}
        is_owner = IsOwner.check_is_owner(user)
        
        if not is_owner and not perms.get('all_centers', False):
            center = serializer.validated_data.get('center')
            if center:
                if user.centers.exists() and center not in user.centers.all() or not user.centers.exists() and hasattr(user, 'center') and center != user.center:
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("You cannot create logs for this center.")
        serializer.save()

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def revenue_report(request):
    user = request.user
    queryset = ServiceLog.objects.all().select_related('staff', 'center', 'invoice')
    role = getattr(user, 'role', None)
    is_owner = IsOwner.check_is_owner(user)
    perms = getattr(role, 'permissions', {}) or {}
    if not is_owner and not perms.get('all_centers', False):
        if user.centers.exists():
            queryset = queryset.filter(center__in=user.centers.all())
        elif hasattr(user, 'center') and user.center:
            queryset = queryset.filter(center=user.center)
            
    center_id = request.query_params.get('center_id')
    if center_id:
        queryset = queryset.filter(center_id=center_id)

    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    if start_date and end_date:
        queryset = queryset.filter(date__range=[start_date, end_date])

    queryset = queryset.filter(Q(invoice__isnull=True) | Q(invoice__status__in=['paid', 'partial']))

    total_revenue = queryset.aggregate(total=Sum('price'))['total'] or 0
    total_services = queryset.count()
    total_staff = queryset.values('staff').distinct().count()

    staff_breakdown = queryset.values('staff__first_name', 'staff__last_name', 'center__display_name').annotate(
        services=Count('id'),
        revenue=Sum('price')
    ).order_by('-revenue')

    breakdown = []
    for item in staff_breakdown:
        name = item['staff__first_name']
        if item['staff__last_name']:
            name += f" {item['staff__last_name']}"
        breakdown.append({
            'staff_name': name,
            'location': item['center__display_name'],
            'services': item['services'],
            'revenue': item['revenue']
        })

    feed = []
    for log in queryset.select_related('staff').order_by('-date', '-time')[:50]:
        staff_name = log.staff.first_name
        if log.staff.last_name:
            staff_name += f" {log.staff.last_name}"
        feed.append({
            'staff_name': staff_name,
            'client_name': log.client_name,
            'service_name': log.service_name,
            'service_type': log.service_type,
            'time': log.time.strftime('%I:%M %p') if log.time else '',
            'price': float(log.price)
        })

    return Response({
        'total_revenue': total_revenue,
        'total_services': total_services,
        'total_staff': total_staff,
        'breakdown': breakdown,
        'activity_feed': feed
    })

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def usage_report(request):
    user = request.user
    queryset = ServiceLog.objects.all().select_related('staff', 'center', 'invoice')
    role = getattr(user, 'role', None)
    is_owner = IsOwner.check_is_owner(user)
    perms = getattr(role, 'permissions', {}) or {}
    if not is_owner and not perms.get('all_centers', False):
        if user.centers.exists():
            queryset = queryset.filter(center__in=user.centers.all())
        elif hasattr(user, 'center') and user.center:
            queryset = queryset.filter(center=user.center)

    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    if start_date and end_date:
        queryset = queryset.filter(date__range=[start_date, end_date])

    queryset = queryset.filter(Q(invoice__isnull=True) | Q(invoice__status__in=['paid', 'partial']))

    services_count = queryset.filter(service_type='Service').count()
    products_count = queryset.filter(service_type='Product').count()
    memberships_count = queryset.filter(service_type='Membership').count()
    packages_count = queryset.filter(service_type='Package').count()

    item_breakdown = queryset.values('service_name', 'service_type').annotate(
        times_used=Count('id'),
        revenue=Sum('price')
    ).order_by('-times_used')

    return Response({
        'services_count': services_count,
        'products_count': products_count,
        'memberships_count': memberships_count,
        'packages_count': packages_count,
        'breakdown': list(item_breakdown)
    })

class StaffConsumptionLogViewSet(viewsets.ModelViewSet):
    serializer_class = StaffConsumptionLogSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        queryset = StaffConsumptionLog.objects.all().select_related('staff', 'center').order_by('-date', '-time')
        role = getattr(user, 'role', None)
        is_owner = IsOwner.check_is_owner(user)
        perms = getattr(role, 'permissions', {}) or {}

        if not is_owner and not perms.get('all_centers', False):
            if user.centers.exists():
                queryset = queryset.filter(center__in=user.centers.all())
            elif hasattr(user, 'center') and user.center:
                queryset = queryset.filter(center=user.center)

        staff_id = self.request.query_params.get('staff_id')
        if staff_id:
            queryset = queryset.filter(staff_id=staff_id)

        center_id = self.request.query_params.get('center_id')
        if center_id:
            queryset = queryset.filter(center_id=center_id)

        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date and end_date:
            queryset = queryset.filter(date__range=[start_date, end_date])

        return queryset


    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def bulk_upload(self, request):
        return Response({'detail': 'Bulk upload for this module is not supported yet.'}, status=400)

    def perform_create(self, serializer):
        user = self.request.user
        perms = getattr(user.role, 'permissions', {}) or {}
        is_owner = IsOwner.check_is_owner(user)
        
        if not is_owner and not perms.get('all_centers', False):
            center = serializer.validated_data.get('center')
            if center:
                if user.centers.exists() and center not in user.centers.all() or not user.centers.exists() and hasattr(user, 'center') and center != user.center:
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("You cannot create logs for this center.")

        payment_method = serializer.validated_data.get('payment_method', 'Points')
        amount = serializer.validated_data.get('amount', 0)
        staff = serializer.validated_data.get('staff')
        service_name = serializer.validated_data.get('service_name')
        
        remaining_money_amount = 0

        if payment_method == 'Points':
            if staff.allocated_points < amount:
                if staff.allocated_points <= 0:
                    serializer.validated_data['payment_method'] = 'Money'
                    payment_method = 'Money'
                else:
                    points_to_deduct = staff.allocated_points
                    remaining_money_amount = amount - points_to_deduct
                    
                    serializer.validated_data['amount'] = points_to_deduct
                    staff.allocated_points = 0
                    staff.save()
            else:
                staff.allocated_points -= amount
                staff.save()

        try:
            from django.db.models import F

            from inventory.models import Product
            product = Product.objects.filter(name__iexact=service_name, center=staff.center).first()
            if product and hasattr(product, 'current_stock'):
                qty = int(serializer.validated_data.get('quantity', 1)) if 'quantity' in serializer.fields else 1
                Product.objects.filter(pk=product.pk).update(current_stock=F('current_stock') - qty)
                # Ensure stock does not go below zero
                Product.objects.filter(pk=product.pk, current_stock__lt=0).update(current_stock=0)
        except Exception as e:
            print(f"[StaffConsumptionLog] Error deducting stock: {e}")

        serializer.save()

        if remaining_money_amount > 0:
            StaffConsumptionLog.objects.create(
                staff=staff,
                center=serializer.validated_data.get('center'),
                service_name=service_name,
                date=serializer.validated_data.get('date'),
                time=serializer.validated_data.get('time'),
                payment_method='Money',
                amount=remaining_money_amount
            )

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def consumption_report(request):
    user = request.user
    queryset = StaffConsumptionLog.objects.all().select_related('staff', 'center', 'product')
    role = getattr(user, 'role', None)
    is_owner = IsOwner.check_is_owner(user)
    perms = getattr(role, 'permissions', {}) or {}
    
    if not is_owner and not perms.get('all_centers', False):
        if user.centers.exists():
            queryset = queryset.filter(center__in=user.centers.all())
        elif hasattr(user, 'center') and user.center:
            queryset = queryset.filter(center=user.center)
            
    center_id = request.query_params.get('center_id')
    if center_id:
        queryset = queryset.filter(center_id=center_id)

    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    if start_date and end_date:
        queryset = queryset.filter(date__range=[start_date, end_date])

    report = queryset.values('staff__first_name', 'staff__last_name', 'service_name', 'payment_method', 'center__display_name').annotate(
        total_amount=Sum('amount'),
        times_used=Count('id')
    ).order_by('-total_amount')

    formatted_report = []
    for item in report:
        name = item['staff__first_name']
        if item['staff__last_name']:
            name += f" {item['staff__last_name']}"
            
        formatted_report.append({
            'staff_name': name,
            'service_name': item['service_name'],
            'payment_method': item['payment_method'],
            'center_name': item['center__display_name'],
            'times_used': item['times_used'],
            'total_amount': item['total_amount']
        })

    return Response(formatted_report)

class StaffTransferViewSet(viewsets.ModelViewSet):
    serializer_class = StaffTransferSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        queryset = StaffTransfer.objects.all().select_related('staff', 'from_center', 'to_center').order_by('-created_at')
        role = getattr(user, 'role', None)
        is_owner = IsOwner.check_is_owner(user)
        perms = getattr(role, 'permissions', {}) or {}

        if not is_owner and not perms.get('all_centers', False):
            if user.centers.exists():
                queryset = queryset.filter(to_center__in=user.centers.all()) | queryset.filter(from_center__in=user.centers.all())
            elif hasattr(user, 'center') and user.center:
                queryset = queryset.filter(to_center=user.center) | queryset.filter(from_center=user.center)

        staff_id = self.request.query_params.get('staff_id')
        if staff_id:
            queryset = queryset.filter(staff_id=staff_id)

        return queryset.distinct()


    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def bulk_upload(self, request):
        return Response({'detail': 'Bulk upload for this module is not supported yet.'}, status=400)

    def perform_create(self, serializer):
        user = self.request.user
        perms = getattr(user.role, 'permissions', {}) or {}
        is_owner = IsOwner.check_is_owner(user)
        
        if not is_owner and not perms.get('all_centers', False):
            from_center = serializer.validated_data.get('from_center')
            if from_center:
                if user.centers.exists() and from_center not in user.centers.all() or not user.centers.exists() and hasattr(user, 'center') and from_center != user.center:
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("You cannot transfer staff from this center.")
                    
        transfer = serializer.save()
        staff = transfer.staff
        staff.center = transfer.to_center
        staff.save()

class StaffToolTrackerViewSet(viewsets.ModelViewSet):
    queryset = StaffToolTracker.objects.all().select_related('staff', 'staff__center').order_by('-created_at')
    serializer_class = StaffToolTrackerSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()

        if hasattr(user, 'role') and user.role and user.role.name.lower() == 'owner' or hasattr(user, 'is_superuser') and user.is_superuser:
            pass
        elif hasattr(user, 'center') and user.center:
            queryset = queryset.filter(staff__center=user.center)

        staff_id = self.request.query_params.get('staff_id')
        if staff_id:
            queryset = queryset.filter(staff_id=staff_id)

        return queryset


    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def bulk_upload(self, request):
        return Response({'detail': 'Bulk upload for this module is not supported yet.'}, status=400)

    def perform_create(self, serializer):
        user = self.request.user
        serializer.save()

class PayrollRecordViewSet(viewsets.ModelViewSet):
    queryset = PayrollRecord.objects.all().select_related('staff', 'center').order_by('-created_at')
    serializer_class = PayrollRecordSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        role = getattr(user, 'role', None)
        is_owner = IsOwner.check_is_owner(user)
        perms = getattr(role, 'permissions', {}) or {}
        
        if not is_owner and not perms.get('all_centers', False):
            if user.centers.exists():
                queryset = queryset.filter(center__in=user.centers.all())
            elif hasattr(user, 'center') and user.center:
                queryset = queryset.filter(center=user.center)
                
        staff_id = self.request.query_params.get('staff_id')
        if staff_id:
            queryset = queryset.filter(staff_id=staff_id)
            
        month = self.request.query_params.get('month')
        year = self.request.query_params.get('year')
        if month:
            queryset = queryset.filter(month=month)
        if year:
            queryset = queryset.filter(year=year)
            
        return queryset


    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def bulk_upload(self, request):
        return Response({'detail': 'Bulk upload for this module is not supported yet.'}, status=400)

    def perform_create(self, serializer):
        user = self.request.user
        perms = getattr(user.role, 'permissions', {}) or {}
        is_owner = IsOwner.check_is_owner(user)
        if not is_owner and not perms.get('all_centers', False):
            center = serializer.validated_data.get('center')
            if center:
                if user.centers.exists() and center not in user.centers.all() or not user.centers.exists() and hasattr(user, 'center') and center != user.center:
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("You cannot create payrolls for this center.")
        serializer.save()

    @action(detail=True, methods=['post'])
    def mark_paid(self, request, pk=None):
        payroll = self.get_object()
        payroll.status = 'Paid'
        payroll.save(update_fields=['status', 'updated_at'])
        return Response(PayrollRecordSerializer(payroll).data)

    @action(detail=True, methods=['post'])
    def lock(self, request, pk=None):
        payroll = self.get_object()
        payroll.status = 'Locked'
        payroll.save(update_fields=['status', 'updated_at'])
        return Response(PayrollRecordSerializer(payroll).data)

# ─────────────────────────────────────────────────────────────────────────────
# Staff App Endpoints — Token-Secured (replaces insecure AllowAny pattern)
# ─────────────────────────────────────────────────────────────────────────────

from django.core import signing as _signing

from pos_backend.permissions import IsOwner


def _generate_staff_token(staff):
    """Generate a signed token for the staff mobile app. Valid 30 days."""
    return _signing.dumps(
        {'staff_id': staff.id, 'pin_hash': str(staff.app_password)[-10:]},
        salt='staff-app-token'
    )

def _verify_staff_token(token):
    """Verify a staff app token. Returns StaffMember or None."""
    try:
        data = _signing.loads(token, salt='staff-app-token', max_age=86400 * 30)
        staff = StaffMember.objects.get(
            id=data['staff_id'],
            is_active=True
        )
        if data.get('pin_hash') != str(staff.app_password)[-10:]:
            return None
        return staff
    except Exception:
        return None

def _get_authenticated_staff(request):
    """Extract and validate X-Staff-Token from request headers. Returns staff or None."""
    token = request.headers.get('X-Staff-Token') or request.query_params.get('_token')
    if not token:
        return None
    return _verify_staff_token(token)


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def staff_app_login(request):
    """Staff app login. Returns staff data + auth_token for subsequent requests."""
    identifier = request.data.get('identifier') or request.data.get('phone')
    pin = request.data.get('pin')

    if not identifier:
        return Response({'error': 'Phone or ID is required'}, status=status.HTTP_400_BAD_REQUEST)
    if not pin:
        return Response({'error': 'PIN is required'}, status=status.HTTP_400_BAD_REQUEST)

    staff = None
    try:
        staff = StaffMember.objects.get(phone=identifier, is_active=True)
    except StaffMember.DoesNotExist:
        try:
            staff = StaffMember.objects.get(id=int(identifier), is_active=True)
        except (StaffMember.DoesNotExist, ValueError):
            pass

    if staff and check_password(str(pin), str(staff.app_password or '')):
        data = StaffMemberSerializer(staff).data
        data['auth_token'] = _generate_staff_token(staff)
        return Response(data)
    elif staff and staff.app_password == str(pin):
        # Fallback for old plaintext passwords during transition
        from django.contrib.auth.hashers import make_password
        staff.app_password = make_password(str(pin))
        staff.save(update_fields=['app_password'])
        data = StaffMemberSerializer(staff).data
        data['auth_token'] = _generate_staff_token(staff)
        return Response(data)

    return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def staff_app_logs(request):
    """Return service logs for the authenticated staff member on a given date."""
    staff = _get_authenticated_staff(request)
    if not staff:
        return Response({'error': 'Unauthorized. Please log in again.'}, status=status.HTTP_401_UNAUTHORIZED)

    target_date_str = request.query_params.get('date')
    from datetime import date, datetime
    if target_date_str:
        try:
            target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
        except ValueError:
            target_date = date.today()
    else:
        target_date = date.today()

    logs = ServiceLog.objects.filter(staff=staff, date=target_date).exclude(invoice__status__in=['cancelled', 'refunded']).order_by('-created_at')
    return Response(ServiceLogSerializer(logs, many=True).data)


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def staff_app_appointments(request):
    """Return appointments for the authenticated staff member on a given date."""
    staff = _get_authenticated_staff(request)
    if not staff:
        return Response({'error': 'Unauthorized. Please log in again.'}, status=status.HTTP_401_UNAUTHORIZED)

    try:
        from datetime import date, datetime

        from appointments.models import Appointment
        from appointments.serializers import AppointmentSerializer

        target_date_str = request.query_params.get('date')
        if target_date_str:
            try:
                target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
            except ValueError:
                target_date = date.today()
        else:
            target_date = date.today()

        appointments = (
            Appointment.objects
            .filter(services__staff=staff, date=target_date)
            .distinct()
            .order_by('date', 'start_time')
        )
        return Response(AppointmentSerializer(appointments, many=True).data)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def staff_app_tools(request):
    """Return tools assigned to the authenticated staff member."""
    staff = _get_authenticated_staff(request)
    if not staff:
        return Response({'error': 'Unauthorized. Please log in again.'}, status=status.HTTP_401_UNAUTHORIZED)

    sync_staff_transfers_and_tools()
    tools = StaffToolTracker.objects.filter(staff=staff).order_by('-created_at')
    return Response(StaffToolTrackerSerializer(tools, many=True).data)


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def staff_app_transfers(request):
    """Return transfers for the authenticated staff member."""
    staff = _get_authenticated_staff(request)
    if not staff:
        return Response({'error': 'Unauthorized. Please log in again.'}, status=status.HTTP_401_UNAUTHORIZED)

    sync_staff_transfers_and_tools()
    transfers = StaffTransfer.objects.filter(staff=staff).order_by('-created_at')
    return Response(StaffTransferSerializer(transfers, many=True).data)


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def staff_app_update_profile(request):
    """Allow staff to update their own Aadhaar number."""
    staff = _get_authenticated_staff(request)
    if not staff:
        return Response({'error': 'Unauthorized. Please log in again.'}, status=status.HTTP_401_UNAUTHORIZED)

    aadhar = request.data.get('aadhar_number')
    if aadhar is not None:
        staff.aadhar_number = aadhar
        staff.save(update_fields=['aadhar_number'])

    return Response({'success': True, 'message': 'Profile updated.'})


