from rest_framework import viewsets
from django.db.models import Count
from .models import Center, Role
from .serializers import CenterSerializer, RoleSerializer
from accounts.access import has_action_permission, has_global_access
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import PermissionDenied
from accounts.permissions import RoleActionPermission

class CenterViewSet(viewsets.ModelViewSet):
    serializer_class = CenterSerializer
    permission_classes = [RoleActionPermission]

    def _require_write(self, action):
        if not has_action_permission(self.request.user, 'admin', 'centers', action):
            raise PermissionDenied('You do not have permission to modify centers.')

    def perform_create(self, serializer):
        self._require_write('create')
        serializer.save()

    def perform_update(self, serializer):
        self._require_write('update')
        serializer.save()

    def perform_destroy(self, instance):
        self._require_write('delete')
        instance.delete()

    def get_queryset(self):
        user = self.request.user
        if not user.is_authenticated:
            return Center.objects.none()
            
        role = getattr(user, 'role', None)
        perms = getattr(role, 'permissions', {}) or {}
        is_owner = getattr(user, 'is_superuser', False) or (role and role.name.lower() == 'owner')
        
        if is_owner or perms.get('all_centers', False):
            qs = Center.objects.all()
        elif user.centers.exists():
            qs = user.centers.all()
        elif hasattr(user, 'center') and user.center:
            qs = Center.objects.filter(id=user.center.id)
        else:
            return Center.objects.none()

        if self.request.GET.get('with_revenue') == 'true':
            from django.db.models import Sum, Q
            from datetime import date
            today = date.today()
            first_day = today.replace(day=1)
            qs = qs.annotate(
                mtd_revenue=Sum(
                    'invoices__total_amount', 
                    filter=Q(invoices__status__in=['paid', 'partial'], invoices__created_at__gte=first_day)
                )
            )
        return qs

class RoleViewSet(viewsets.ModelViewSet):
    queryset = Role.objects.annotate(users_count=Count('customuser', distinct=True)).all()
    serializer_class = RoleSerializer
    permission_classes = [RoleActionPermission]

    def get_queryset(self):
        if not has_action_permission(self.request.user, 'admin', 'roles', 'read'):
            return self.queryset.none()
        return self.queryset

    def _require_write(self, action):
        if not has_action_permission(self.request.user, 'admin', 'roles', action):
            raise PermissionDenied('You do not have permission to modify roles.')

    def perform_create(self, serializer):
        self._require_write('create')
        serializer.save()

    def perform_update(self, serializer):
        self._require_write('update')
        serializer.save()

    def perform_destroy(self, instance):
        self._require_write('delete')
        instance.delete()

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.db.models import Sum, Count
from billing.models import Invoice
from appointments.models import Appointment
from clients.models import Client
from datetime import datetime, timedelta

@api_view(['GET'])
@permission_classes([RoleActionPermission])
def dashboard_view(request):
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    center_id = request.GET.get('center_id')
    
    invoices = Invoice.objects.filter(status__in=['paid', 'partial'])
    appointments = Appointment.objects.all()
    clients = Client.objects.all()

    user = request.user
    perms = getattr(user.role, 'permissions', {}) if getattr(user, 'role', None) else {}
    is_owner = getattr(user, 'is_superuser', False) or (getattr(user, 'role', None) and user.role.name.lower() == 'owner')
    if not is_owner and not perms.get('all_centers', False):
        if user.centers.exists():
            invoices = invoices.filter(center__in=user.centers.all())
            appointments = appointments.filter(center__in=user.centers.all())
            clients = clients.filter(center__in=user.centers.all())
        elif hasattr(user, 'center') and user.center:
            invoices = invoices.filter(center=user.center)
            appointments = appointments.filter(center=user.center)
            clients = clients.filter(center=user.center)
    
    if center_id and center_id != 'null':
        invoices = invoices.filter(center_id=center_id)
        appointments = appointments.filter(center_id=center_id)
        clients = clients.filter(center_id=center_id)
        
    if start_date_str and end_date_str:
        try:
            end_date_obj = datetime.strptime(end_date_str, '%Y-%m-%d').date() + timedelta(days=1)
            invoices = invoices.filter(created_at__gte=start_date_str, created_at__lt=end_date_obj)
            appointments = appointments.filter(date__gte=start_date_str, date__lte=end_date_str)
            clients = clients.filter(created_at__gte=start_date_str, created_at__lt=end_date_obj)
        except ValueError:
            pass

    # Aggregate using database efficiently
    total_revenue = invoices.aggregate(total=Sum('total_amount'))['total'] or 0
    total_appointments = appointments.count()
    total_clients = clients.count()
    
    from django.db.models import DateField
    from django.db.models.functions import Cast
    
    # Generate daily breakdown directly efficiently
    daily_rev = invoices.annotate(day=Cast('created_at', DateField())).values('day').annotate(revenue=Sum('total_amount')).order_by('day')
    revenue_by_day = [{'day': item['day'].strftime('%Y-%m-%d') if hasattr(item['day'], 'strftime') else str(item['day']), 'revenue': float(item['revenue'] or 0)} for item in daily_rev if item['day']]
    
    daily_appts = appointments.values('date').annotate(count=Count('id')).order_by('date')
    appointments_by_day = [{'day': item['date'].strftime('%Y-%m-%d') if isinstance(item['date'], datetime) else str(item['date']), 'count': item['count']} for item in daily_appts if item['date']]
    
    appt_status_counts = appointments.values('status').annotate(count=Count('id'))
    appt_by_status = [{'status': item['status'], 'count': item['count']} for item in appt_status_counts]

    try:
        from billing.models import InvoiceItem
        service_count = InvoiceItem.objects.filter(invoice__in=invoices, content_type__model='servicemaster').aggregate(t=Sum('quantity'))['t'] or 0
        product_count = InvoiceItem.objects.filter(invoice__in=invoices, content_type__model='product').aggregate(t=Sum('quantity'))['t'] or 0
    except Exception:
        service_count = 0
        product_count = 0

    service_type_breakdown = [
        {'service_type': 'Service', 'count': service_count},
        {'service_type': 'Product', 'count': product_count}
    ]

# New vs Repeat Clients Logic
    try:
        new_client_count = 0
        repeat_client_count = 0
        
        # Get all distinct clients who had an invoice in this period
        visiting_client_ids = invoices.exclude(client__isnull=True).values_list('client_id', flat=True).distinct()
        
        if start_date_str:
            sd_obj = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            # Out of the visitors, how many were created on or after the start date?
            new_client_count = Client.objects.filter(id__in=visiting_client_ids, created_at__date__gte=sd_obj).count()
            repeat_client_count = len(visiting_client_ids) - new_client_count
        else:
            # If no date range is given, technically all are historical/repeat or we can just count all as repeat
            repeat_client_count = len(visiting_client_ids)
    except Exception as e:
        new_client_count = 0
        repeat_client_count = 0

    new_vs_repeat = [
        {'label': 'New', 'count': new_client_count},
        {'label': 'Repeat', 'count': repeat_client_count}
    ]

    try:
        from billing.models import InvoiceItem
        items = InvoiceItem.objects.filter(invoice__in=invoices, content_type__model='servicemaster')

        # Service-wise revenue — group by description (item_name does not exist on InvoiceItem)
        service_revenue = items.values('description').annotate(
            revenue=Sum('total_price')
        ).order_by('-revenue')

        service_revenue_breakdown = [{'name': s['description'] or 'Unknown', 'revenue': float(s['revenue'] or 0)} for s in service_revenue]
        top_5_services = service_revenue_breakdown[:5]
    except Exception as e:
        service_revenue_breakdown = []
        top_5_services = []

    footfall = invoices.count()
    avg_spend = (total_revenue / footfall) if footfall > 0 else 0

    return Response({
        'total_revenue': total_revenue,
        'total_appointments': total_appointments,
        'total_clients': total_clients,
        'footfall': footfall,
        'avg_spend': round(float(avg_spend), 2),
        'revenue_by_day': revenue_by_day,
        'service_type_breakdown': service_type_breakdown,
        'service_revenue_breakdown': service_revenue_breakdown,
        'top_5_services': top_5_services,
        'new_vs_repeat': new_vs_repeat,
        'appointments_by_day': appointments_by_day,
        'appt_by_status': appt_by_status,
        'visits_by_day': appointments_by_day,
    })


# ─────────────────────────────────────────────────────────
#  BULK IMPORT CENTRES FROM EXCEL
# ─────────────────────────────────────────────────────────
@api_view(['POST'])
@permission_classes([RoleActionPermission])
def bulk_import_centers(request):
    import io
    try:
        import openpyxl
    except ImportError:
        return Response({'error': 'openpyxl not installed. Run: pip install openpyxl'}, status=400)

    user = request.user
    role = getattr(user, 'role', None)
    perms = getattr(role, 'permissions', {}) or {}
    is_owner = getattr(user, 'is_superuser', False) or (role and role.name.lower() == 'owner')
    if not is_owner and not perms.get('all_centers', False):
        return Response({'error': 'Permission denied. Only owners can bulk import centres.'}, status=403)

    uploaded_file = request.FILES.get('file')
    if not uploaded_file:
        return Response({'error': 'No file uploaded. Send the Excel as multipart key "file".'}, status=400)

    COLUMN_MAP = {
        'center name': 'center_name', 'centre name': 'center_name',
        'name': 'center_name', 'salon name': 'center_name', 'legal name': 'center_name',
        'display name': 'display_name', 'short name': 'display_name',
        'address': 'address', 'location': 'address', 'full address': 'address',
        'region': 'region', 'city': 'region', 'state': 'region',
        'phone': 'phone', 'phone 1': 'phone', 'mobile': 'phone', 'contact': 'phone',
        'landline 1': 'landline_1', 'landline1': 'landline_1', 'phone 2': 'landline_1', 'alternate phone': 'landline_1',
        'landline 2': 'landline_2', 'landline2': 'landline_2',
        'center email': 'center_email', 'centre email': 'center_email', 'email': 'center_email',
        'gst': 'gst_number', 'gst number': 'gst_number', 'gst no': 'gst_number', 'gstin': 'gst_number',
        'pan': 'pan_number', 'pan number': 'pan_number', 'pan no': 'pan_number',
        'monthly target': 'monthly_target', 'target': 'monthly_target',
        'owner name': 'owner_name', 'owner': 'owner_name', 'owner 1': 'owner_name',
        'owner phone': 'owner_phone', 'owner mobile': 'owner_phone',
        'owner email': 'owner_email_1', 'owner email 1': 'owner_email_1',
        'owner name 2': 'owner_name_2', 'owner phone 2': 'owner_phone_2', 'owner email 2': 'owner_email_2',
        'owner name 3': 'owner_name_3', 'owner phone 3': 'owner_phone_3', 'owner email 3': 'owner_email_3',
        'accountant': 'accountant_name_1', 'accountant name': 'accountant_name_1', 'accountant name 1': 'accountant_name_1',
        'accountant phone': 'accountant_phone_1', 'accountant email': 'accountant_email_1',
        'accountant name 2': 'accountant_name_2', 'accountant phone 2': 'accountant_phone_2', 'accountant email 2': 'accountant_email_2',
        'launched on': 'launched_on', 'opening date': 'launched_on', 'start date': 'launched_on', 'date': 'launched_on',
        'salon/studio': 'display_name', 'salon / studio': 'display_name', 'type': None,
    }

    try:
        wb = openpyxl.load_workbook(io.BytesIO(uploaded_file.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        return Response({'error': f'Could not read Excel file: {str(e)}'}, status=400)

    # Fix mapping based on user's screenshot
    COLUMN_MAP['location'] = 'center_name' # In their sheet, Location is the Center Name
    COLUMN_MAP['state'] = 'region'
    COLUMN_MAP['date'] = 'launched_on'

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return Response({'error': 'Excel file is empty.'}, status=400)

    # Find the header row (scan first 10 rows)
    header_row_idx = 0
    col_map = {}
    header = []
    
    for i in range(min(10, len(rows))):
        current_header = [str(h).strip().lower() if h is not None else '' for h in rows[i]]
        current_map = {}
        for idx, h in enumerate(current_header):
            field = COLUMN_MAP.get(h)
            if field:
                current_map[idx] = field
        if len(current_map) > len(col_map):
            col_map = current_map
            header = current_header
            header_row_idx = i

    if not col_map:
        return Response({
            'error': 'No recognised column headers found.',
            'hint': 'Use headers like: Center Name, Display Name, Address, Region, Phone, GST Number, Owner Name, etc.',
            'found_headers': [str(h) for h in (rows[0] if len(rows) > 0 else [])],
        }, status=400)

    created_list, updated_list, skipped_list, errors_list = [], [], [], []
    import datetime as dt_mod

    for row_num, row in enumerate(rows[header_row_idx+1:], start=header_row_idx+2):
        data = {}
        for col_idx, field_name in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None and str(val).strip():
                data[field_name] = str(val).strip()

        center_name = data.get('center_name', '').strip()
        if not center_name:
            skipped_list.append({'row': row_num, 'reason': 'Missing center name'})
            continue

        if 'launched_on' in data:
            raw = data['launched_on']
            parsed = None
            for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y', '%d-%b-%Y'):
                try:
                    parsed = dt_mod.datetime.strptime(str(raw), fmt).strftime('%Y-%m-%d')
                    break
                except ValueError:
                    continue
            if parsed:
                data['launched_on'] = parsed
            else:
                data.pop('launched_on', None)

        try:
            display_name = data.get('display_name', '').strip()
            
            if display_name:
                existing = Center.objects.filter(center_name__iexact=center_name, display_name__iexact=display_name).first()
            else:
                existing = Center.objects.filter(center_name__iexact=center_name).first()
                
            if existing:
                for field, value in data.items():
                    if field != 'center_name':
                        setattr(existing, field, value)
                existing.save()
                updated_list.append({'row': row_num, 'center_name': center_name})
            else:
                center = Center(center_name=center_name)
                for field, value in data.items():
                    if field != 'center_name':
                        setattr(center, field, value)
                center.save()
                created_list.append({'row': row_num, 'center_name': center_name})
        except Exception as e:
            errors_list.append({'row': row_num, 'center_name': center_name, 'error': str(e)})

    return Response({
        'success': True,
        'summary': {
            'total_rows': len(rows) - 1,
            'created': len(created_list),
            'updated': len(updated_list),
            'skipped': len(skipped_list),
            'errors': len(errors_list),
        },
        'created': created_list,
        'updated': updated_list,
        'skipped': skipped_list,
        'errors': errors_list,
    })

@api_view(['GET'])
@permission_classes([RoleActionPermission])
def bulk_import_template(request):
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return Response({'error': 'openpyxl not installed.'}, status=400)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Centres Import"
    
    headers = [
        'Center Name', 'Display Name', 'Address', 'Region', 'Phone', 'Landline 1',
        'Landline 2', 'Center Email',
        'GST Number', 'PAN Number', 'Monthly Target', 'Owner Name', 'Owner Phone',
        'Owner Email', 'Owner Name 2', 'Owner Phone 2', 'Owner Email 2', 
        'Owner Name 3', 'Owner Phone 3', 'Owner Email 3',
        'Accountant Name', 'Accountant Phone', 'Accountant Email',
        'Accountant Name 2', 'Accountant Phone 2', 'Accountant Email 2',
        'Launched On'
    ]
    
    ws.append(headers)
    
    # Styling
    header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # Adjust width
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
        
    help_text = [
        'Required', 'Optional', 'Optional', 'Optional', 'Optional', 'Optional',
        'Optional', 'Optional',
        'Optional', 'Optional', 'Optional Number', 'Optional', 'Optional',
        'Optional', 'Optional', 'Optional', 'Optional',
        'Optional', 'Optional', 'Optional',
        'Optional', 'Optional', 'Optional',
        'Optional', 'Optional', 'Optional',
        'YYYY-MM-DD or DD-MM-YYYY'
    ]
    ws.append(help_text)
    
    help_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
    help_font = Font(color="15803D", italic=True)
    for cell in ws[2]:
        cell.fill = help_fill
        cell.font = help_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    response_io = io.BytesIO()
    wb.save(response_io)
    response_io.seek(0)
    
    from django.http import HttpResponse
    response = HttpResponse(response_io.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="centres_import_template.xlsx"'
    return response
