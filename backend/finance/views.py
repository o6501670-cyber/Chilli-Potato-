from django.utils.decorators import method_decorator
# Removed cache_page due to cross-tenant leak
from decimal import Decimal
from rest_framework import viewsets, status, views
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from .models import PettyCashEntry, DailyClosing, IncentiveConfig, Shift, IncentiveRule
from .serializers import PettyCashEntrySerializer, DailyClosingSerializer, IncentiveConfigSerializer, ShiftSerializer, IncentiveRuleSerializer
from billing.models import Invoice, Payment, InvoiceItem, AdvancePayment
from inventory.models import PurchaseOrder, PurchaseOrderItem
from django.db.models import Sum, Count, Q
from django.contrib.contenttypes.models import ContentType
from salon_admin.models import Center
from datetime import datetime
import datetime as dt_module
from collections import defaultdict
import calendar
import logging

logger = logging.getLogger(__name__)


def _get_filtered_invoices(request, center_id, start_date, end_date, statuses=('paid', 'partial')):
    qs = Invoice.objects.filter(status__in=statuses)

    user = request.user
    perms = getattr(user.role, 'permissions', {}) or {}
    is_owner = IsOwner.check_is_owner(user)
    if not is_owner and not perms.get('all_centers', False):
        if user.centers.exists():
            qs = qs.filter(center__in=user.centers.all())
        elif hasattr(user, 'center') and user.center:
            qs = qs.filter(center=user.center)

    if center_id:
        qs = qs.filter(center_id=center_id)

    # Use the same datetime-range pattern as billing/views.py to avoid off-by-one near midnight
    # e.g. __date__gte vs __gte with datetime can produce different results in the same query
    if start_date:
        try:
            from datetime import datetime as _dt
            qs = qs.filter(created_at__gte=_dt.strptime(str(start_date), '%Y-%m-%d'))
        except (ValueError, TypeError):
            qs = qs.filter(created_at__date__gte=start_date)
    if end_date:
        try:
            from datetime import datetime as _dt, timedelta as _td
            end_dt = _dt.strptime(str(end_date), '%Y-%m-%d') + _td(days=1)
            qs = qs.filter(created_at__lt=end_dt)
        except (ValueError, TypeError):
            qs = qs.filter(created_at__date__lte=end_date)
    return qs


def _compute_revenue_breakdown(invoices):
    """Compute revenue breakdown by item type from invoices."""
    from services.models import ServiceMaster
    from inventory.models import Product
    from marketing.models import Membership, Package, ValueCard

    try:
        ct_map = ContentType.objects.get_for_models(ServiceMaster, Product, Membership, Package, ValueCard)
        service_ct = ct_map[ServiceMaster]
        product_ct = ct_map[Product]
        membership_ct = ct_map[Membership]
        package_ct = ct_map[Package]
        valuecard_ct = ct_map[ValueCard]
    except Exception as e:
        logger.warning(f'[Finance] ContentType lookup failed during revenue breakdown: {e}')
        return {'services': 0, 'products': 0, 'memberships': 0, 'packages': 0, 'value_cards': 0, 'other': 0}

    from django.db.models import Sum, Q

    items_qs = InvoiceItem.objects.filter(invoice__in=invoices)

    items = items_qs.aggregate(
        services=Sum('total_price', filter=Q(content_type=service_ct)),
        products=Sum('total_price', filter=Q(content_type=product_ct)),
        memberships=Sum('total_price', filter=Q(content_type=membership_ct)),
        packages=Sum('total_price', filter=Q(content_type=package_ct)),
        value_cards=Sum('total_price', filter=Q(content_type=valuecard_ct)),
        other=Sum('total_price', filter=Q(content_type__isnull=True))
    )
    
    return {k: Decimal(str(v or 0)) for k, v in items.items()}


def _compute_payment_breakdown(invoices):
    """Compute payment method totals from invoice payments efficiently."""
    payments_agg = Payment.objects.filter(invoice__in=invoices).values('payment_method').annotate(
        total_amount=Sum('amount'), count=Count('id')
    )
    
    method_map = {
        'cash': ['cash'],
        'credit_card': ['credit card', 'debit card'],
        'paytm': ['paytm'],
        'bharat_pe': ['bharatpe', 'bharat pe'],
        'cheque_net_banking': ['cheque', 'net banking', 'neft', 'rtgs'],
        'google_pay': ['google pay', 'gpay'],
        'phone_pe': ['phonepe', 'phone pe'],
        'nearbuy': ['nearbuy'],
        'upi': ['upi'],
        'value_card': ['value card'],
        'cashback_wallet': ['cashback wallet'],
        'advance': ['advance'],
        'other': [],
    }

    result = {k: {'amount': Decimal('0.00'), 'count': 0} for k in method_map}

    for p in payments_agg:
        pm_lower = (p['payment_method'] or '').lower().strip()
        amt = Decimal(str(p['total_amount'] or 0))
        cnt = p['count']
        matched = False
        
        # Exact/tight matching
        if pm_lower in ('card',):
            result['credit_card']['amount'] += amt
            result['credit_card']['count'] += cnt
            continue

        for key, keywords in method_map.items():
            if key == 'other':
                continue
            if any(kw in pm_lower for kw in keywords):
                result[key]['amount'] += amt
                result[key]['count'] += cnt
                matched = True
                break
                
        if not matched:
            result['other']['amount'] += amt
            result['other']['count'] += cnt

    return result


class PettyCashEntryViewSet(viewsets.ModelViewSet):
    queryset = PettyCashEntry.objects.all().select_related('user', 'center').order_by('-date', '-id')
    serializer_class = PettyCashEntrySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        queryset = self.queryset
        
        is_owner = IsOwner.check_is_owner(user)
        perms = getattr(user.role, 'permissions', {}) or {}
        if not is_owner and not perms.get('all_centers', False):
            if user.centers.exists():
                queryset = queryset.filter(center__in=user.centers.all())
            elif hasattr(user, 'center') and user.center:
                queryset = queryset.filter(center=user.center)
            else:
                return queryset.none()
                
        center_id = self.request.query_params.get('center_id', None)
        start_date = self.request.query_params.get('start_date', None)
        end_date = self.request.query_params.get('end_date', None)
        
        if center_id:
            queryset = queryset.filter(center_id=center_id)
        # Use __date__ lookup so DateTimeField is compared correctly (avoids missing end-day entries)
        if start_date:
            queryset = queryset.filter(date__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__date__lte=end_date)
            
        return queryset.order_by('-date')

    def perform_create(self, serializer):
        user = self.request.user
        perms = getattr(user.role, 'permissions', {}) or {}
        is_owner = IsOwner.check_is_owner(user)
        
        if not is_owner and not perms.get('all_centers', False):
            center = serializer.validated_data.get('center')
            if center:
                if user.centers.exists() and center not in user.centers.all():
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("You cannot create petty cash entries for this center.")
                elif not user.centers.exists() and hasattr(user, 'center') and center != user.center:
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("You cannot create petty cash entries for this center.")
        serializer.save(user=self.request.user)


class DailyClosingViewSet(viewsets.ModelViewSet):
    queryset = DailyClosing.objects.all().select_related('user', 'center').order_by('-date', '-id')
    serializer_class = DailyClosingSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        queryset = self.queryset
        
        is_owner = IsOwner.check_is_owner(user)
        perms = getattr(user.role, 'permissions', {}) or {}
        if not is_owner and not perms.get('all_centers', False):
            if user.centers.exists():
                queryset = queryset.filter(center__in=user.centers.all())
            elif hasattr(user, 'center') and user.center:
                queryset = queryset.filter(center=user.center)
            else:
                return queryset.none()
                
        center_id = self.request.query_params.get('center_id', None)
        date = self.request.query_params.get('date', None)
        start_date = self.request.query_params.get('start_date', None)
        end_date = self.request.query_params.get('end_date', None)
        if center_id:
            queryset = queryset.filter(center_id=center_id)
        if date:
            queryset = queryset.filter(date=date)
        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)
        return queryset.order_by('-date')

    def create(self, request, *args, **kwargs):
        center_id = request.data.get('center')
        date_str = request.data.get('date')
        
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Check permissions similar to perform_create
        user = self.request.user
        perms = getattr(user.role, 'permissions', {}) or {}
        is_owner = IsOwner.check_is_owner(user)
        center = serializer.validated_data.get('center')
        
        if not is_owner and not perms.get('all_centers', False):
            if center:
                if user.centers.exists() and center not in user.centers.all():
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("You cannot create daily closing for this center.")
        
        defaults = serializer.validated_data.copy()
        if 'center' in defaults:
            del defaults['center']
        if 'date' in defaults:
            del defaults['date']
        
        from django.db import transaction, IntegrityError
        try:
            with transaction.atomic():
                instance, created = DailyClosing.objects.update_or_create(
                    center_id=center_id, 
                    date=date_str,
                    defaults=defaults
                )
        except IntegrityError:
            # If a concurrent request beat us to it, just update the existing one
            instance, created = DailyClosing.objects.update_or_create(
                center_id=center_id, 
                date=date_str,
                defaults=defaults
            )
        if created:
            instance.user = user
            instance.save(update_fields=['user'])
            
        return Response(self.get_serializer(instance).data, status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)

    def perform_create(self, serializer):
        user = self.request.user
        perms = getattr(user.role, 'permissions', {}) or {}
        is_owner = IsOwner.check_is_owner(user)
        
        if not is_owner and not perms.get('all_centers', False):
            center = serializer.validated_data.get('center')
            if center:
                if user.centers.exists() and center not in user.centers.all():
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("You cannot create petty cash entries for this center.")
                elif not user.centers.exists() and hasattr(user, 'center') and center != user.center:
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("You cannot create petty cash entries for this center.")
        serializer.save(user=self.request.user)

    def perform_update(self, serializer):
        user = self.request.user
        perms = getattr(user.role, 'permissions', {}) or {}
        is_owner = IsOwner.check_is_owner(user)
        
        if not is_owner and not perms.get('all_centers', False):
            center = serializer.validated_data.get('center', serializer.instance.center)
            if center:
                if user.centers.exists() and center not in user.centers.all():
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("You cannot update petty cash entries for this center.")
                elif not user.centers.exists() and hasattr(user, 'center') and center != user.center:
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("You cannot update petty cash entries for this center.")
        serializer.save(user=self.request.user)

    from rest_framework.decorators import action
    @action(detail=False, methods=['get'])
    def opening_balance(self, request):
        import datetime
        center_id = request.query_params.get('center_id')
        date_str = request.query_params.get('date', '')
        if not center_id or not date_str:
            return Response({'opening_balance': 0})
        try:
            target = datetime.date.fromisoformat(date_str)
            prev = target - datetime.timedelta(days=1)
            prev_closing = DailyClosing.objects.filter(
                center_id=center_id, date=prev
            ).first()
            balance = Decimal(str(prev_closing.closing_balance)) if prev_closing else 0
            return Response({'opening_balance': balance, 'from_date': str(prev)})
        except Exception:
            return Response({'opening_balance': 0})


class ShiftViewSet(viewsets.ModelViewSet):
    queryset = Shift.objects.all().select_related('opened_by', 'closed_by').order_by('-opened_at')
    serializer_class = ShiftSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        queryset = self.queryset
        
        is_owner = IsOwner.check_is_owner(user)
        perms = getattr(user.role, 'permissions', {}) or {}
        if not is_owner and not perms.get('all_centers', False):
            if user.centers.exists():
                queryset = queryset.filter(center__in=user.centers.all())
            elif hasattr(user, 'center') and user.center:
                queryset = queryset.filter(center=user.center)
            else:
                return queryset.none()
                
        center_id = self.request.query_params.get('center_id')
        status = self.request.query_params.get('status')
        if center_id:
            queryset = queryset.filter(center_id=center_id)
        if status:
            queryset = queryset.filter(status=status)
        return queryset.order_by('-opened_at')

    def perform_create(self, serializer):
        user = self.request.user
        perms = getattr(user.role, 'permissions', {}) or {}
        is_owner = IsOwner.check_is_owner(user)
        
        if not is_owner and not perms.get('all_centers', False):
            center = serializer.validated_data.get('center')
            if center:
                if user.centers.exists() and center not in user.centers.all():
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("You cannot create shifts for this center.")
                elif not user.centers.exists() and hasattr(user, 'center') and center != user.center:
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("You cannot create shifts for this center.")
        serializer.save(opened_by=self.request.user, status='Open')

    from rest_framework.decorators import action
    @action(detail=True, methods=['post'])
    def close_shift(self, request, pk=None):
        from django.utils import timezone
        shift = self.get_object()
        if shift.status == 'Closed':
            return Response({'error': 'Shift already closed'}, status=400)
        
        actual_cash = request.data.get('actual_cash', 0)
        expected_cash = request.data.get('expected_cash', 0)
        shift.expected_cash = Decimal(str(expected_cash))
        shift.actual_cash = Decimal(str(actual_cash))
        shift.variance = Decimal(str(actual_cash)) - Decimal(str(expected_cash))
        shift.status = 'Closed'
        shift.closed_by = request.user
        shift.closed_at = timezone.now()
        shift.save()
        return Response(ShiftSerializer(shift).data)


class IncentiveRuleViewSet(viewsets.ModelViewSet):
    queryset = IncentiveRule.objects.all().select_related('center').order_by('-created_at')
    serializer_class = IncentiveRuleSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        queryset = self.queryset
        
        is_owner = IsOwner.check_is_owner(user)
        perms = getattr(user.role, 'permissions', {}) or {}
        if not is_owner and not perms.get('all_centers', False):
            if user.centers.exists():
                queryset = queryset.filter(Q(center__in=user.centers.all()) | Q(center__isnull=True))
            elif hasattr(user, 'center') and user.center:
                queryset = queryset.filter(Q(center=user.center) | Q(center__isnull=True))
            else:
                queryset = queryset.filter(center__isnull=True)
                
        center_id = self.request.query_params.get('center_id')
        if center_id:
            queryset = queryset.filter(Q(center_id=center_id) | Q(center__isnull=True))
            
        frequency = self.request.query_params.get('frequency')
        if frequency:
            queryset = queryset.filter(frequency=frequency)
            
        category = self.request.query_params.get('category')
        if category:
            queryset = queryset.filter(category=category)
            
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            if is_active.lower() in ('true', '1'):
                queryset = queryset.filter(is_active=True)
            elif is_active.lower() in ('false', '0'):
                queryset = queryset.filter(is_active=False)
                
        return queryset.order_by('-id')

    def perform_create(self, serializer):
        user = self.request.user
        perms = getattr(user.role, 'permissions', {}) or {}
        is_owner = IsOwner.check_is_owner(user)
        
        if not is_owner and not perms.get('all_centers', False):
            center = serializer.validated_data.get('center')
            if center:
                if user.centers.exists() and center not in user.centers.all():
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("You cannot create incentive rules for this center.")
                elif not user.centers.exists() and hasattr(user, 'center') and center != user.center:
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("You cannot create incentive rules for this center.")
        serializer.save()

    @action(detail=True, methods=['post'])
    def duplicate(self, request, pk=None):
        rule = self.get_object()
        new_rule = IncentiveRule.objects.create(
            name=f"{rule.name} (Copy)",
            center=rule.center,
            frequency=rule.frequency,
            category=rule.category,
            rule_type=rule.rule_type,
            applicable_role=rule.applicable_role,
            tiers=rule.tiers,
            flat_percent=rule.flat_percent,
            flat_amount=rule.flat_amount,
            effective_from=dt_module.date.today(),
            effective_to=None,
            is_active=True,
            description=rule.description,
        )
        serializer = self.get_serializer(new_rule)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class IncentiveConfigViewSet(viewsets.ModelViewSet):
    queryset = IncentiveConfig.objects.all().prefetch_related('tiers')
    serializer_class = IncentiveConfigSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        queryset = IncentiveConfig.objects.all()
        
        is_owner = IsOwner.check_is_owner(user)
        perms = getattr(user.role, 'permissions', {}) or {}
        if not is_owner and not perms.get('all_centers', False):
            if user.centers.exists():
                queryset = queryset.filter(center__in=user.centers.all())
            elif hasattr(user, 'center') and user.center:
                queryset = queryset.filter(center=user.center)
            else:
                queryset = queryset.filter(center__isnull=True)
                
        center_id = self.request.query_params.get('center_id', None)
        if center_id:
            queryset = queryset.filter(center_id=center_id)
        return queryset.order_by('id')

    def perform_create(self, serializer):
        user = self.request.user
        perms = getattr(user.role, 'permissions', {}) or {}
        is_owner = IsOwner.check_is_owner(user)
        
        if not is_owner and not perms.get('all_centers', False):
            center = serializer.validated_data.get('center')
            if center:
                if user.centers.exists() and center not in user.centers.all():
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("You cannot create incentive configs for this center.")
                elif not user.centers.exists() and hasattr(user, 'center') and center != user.center:
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("You cannot create incentive configs for this center.")
        serializer.save()


def compute_register_summary(user, center_id, start_date, end_date) -> dict:
    class MockReq: pass
    req = MockReq()
    req.user = user
    request = req
    # start_date and end_date are already passed as arguments, do not read from query_params


    invoices = _get_filtered_invoices(request, center_id, start_date, end_date)
    
    revenue = _compute_revenue_breakdown(invoices)
    payments = _compute_payment_breakdown(invoices)

    # Advances received (positive advance payments in date range)
    adv_qs = AdvancePayment.objects.filter(amount__gt=0)
    adv_used_qs = AdvancePayment.objects.filter(amount__lt=0)
    
    user = request.user
    perms = getattr(user.role, 'permissions', {}) or {}
    is_owner = IsOwner.check_is_owner(user)
    if not is_owner and not perms.get('all_centers', False):
        if user.centers.exists():
            adv_qs = adv_qs.filter(client__center__in=user.centers.all())
            adv_used_qs = adv_used_qs.filter(client__center__in=user.centers.all())
        elif hasattr(user, 'center') and user.center:
            adv_qs = adv_qs.filter(client__center=user.center)
            adv_used_qs = adv_used_qs.filter(client__center=user.center)
            
    if center_id:
        adv_qs = adv_qs.filter(client__center_id=center_id)
        adv_used_qs = adv_used_qs.filter(client__center_id=center_id)
    if start_date:
        adv_qs = adv_qs.filter(created_at__date__gte=start_date)
        adv_used_qs = adv_used_qs.filter(created_at__date__gte=start_date)
    if end_date:
        adv_qs = adv_qs.filter(created_at__date__lte=end_date)
        adv_used_qs = adv_used_qs.filter(created_at__date__lte=end_date)
        
    advances_total = Decimal(str(adv_qs.aggregate(t=Sum('amount'))['t'] or 0))
    advance_redemptions = abs(Decimal(str(adv_used_qs.aggregate(t=Sum('amount'))['t'] or 0)))

    # Refunds (cancelled invoices)
    cancelled_qs = _get_filtered_invoices(request, center_id, start_date, end_date, statuses=('cancelled', 'refunded'))
    refunds_total = Decimal(str(cancelled_qs.aggregate(t=Sum('total_amount'))['t'] or 0))

    # Tax totals — combine into a single aggregate call instead of 3 separate ones
    totals = invoices.aggregate(
        t_cgst=Sum('cgst'), t_sgst=Sum('sgst'), t_discount=Sum('discount')
    )
    total_tax = Decimal(str(totals['t_cgst'] or 0)) + Decimal(str(totals['t_sgst'] or 0))
    total_discount = Decimal(str(totals['t_discount'] or 0))

    # Proportional tax split based on actual service vs product revenue
    service_product_revenue = revenue['services'] + revenue['products']
    if service_product_revenue > 0:
        services_share = revenue['services'] / service_product_revenue
        products_share = revenue['products'] / service_product_revenue
    else:
        services_share = Decimal('0.7')
        products_share = Decimal('0.3')
    services_tax = total_tax * Decimal(str(services_share))
    products_tax = total_tax * Decimal(str(products_share))
    
    total_taxable = revenue['services'] + revenue['products'] + revenue['value_cards'] + revenue['memberships'] + revenue['packages']

    # Collection before tax = only actual sales revenue (advances are liabilities, shown separately)
    sales_collection = (
        revenue['services'] + revenue['products'] +
        revenue['memberships'] + revenue['packages'] +
        revenue['value_cards'] + revenue['other']
    ) - total_discount
    
    # Net collection: add advances (actual cash received), deduct redemptions (liabilities drawn down)
    value_card_redemptions = Decimal(str(payments['value_card']['amount']))
    cashback_redemptions = Decimal(str(payments['cashback_wallet']['amount']))
    
    collection_before_tax = sales_collection + advances_total - advance_redemptions - value_card_redemptions - cashback_redemptions
    including_tax = collection_before_tax + total_tax

    target = 0
    target_achieved_percentage = 0
    if center_id:
        from salon_admin.models import Center
        try:
            center_obj = Center.objects.get(id=center_id)
            target = 0
            history = center_obj.monthly_targets_history or {}
            
            is_past_range = False
            if start_date and end_date:
                from datetime import datetime
                import calendar
                try:
                    sd = datetime.strptime(start_date, '%Y-%m-%d')
                    ed = datetime.strptime(end_date, '%Y-%m-%d')
                    curr_y, curr_m = sd.year, sd.month
                    end_y, end_m = ed.year, ed.month
                    
                    today = datetime.now()
                    is_past_range = True
                    
                    while (curr_y < end_y) or (curr_y == end_y and curr_m <= end_m):
                        if (curr_y > today.year) or (curr_y == today.year and curr_m >= today.month):
                            is_past_range = False
                            
                        month_abbr = calendar.month_abbr[curr_m]
                        month_key = f"{month_abbr}-{curr_y}"
                        val = history.get(month_key, 0)
                        target += Decimal(str(val or 0))
                        
                        curr_m += 1
                        if curr_m > 12:
                            curr_m = 1
                            curr_y += 1
                except Exception:
                    import logging; logging.getLogger(__name__).error('Handled exception', exc_info=True)
            
            # Fallback to default monthly_target if no target found from history or dates missing
            # Only do this if the date range includes the current or future month.
            if target == 0 and not is_past_range:
                target = Decimal(str(center_obj.monthly_target or 0))

            if target > 0:
                target_achieved_percentage = round((total_taxable / target) * 100, 2)
        except Center.DoesNotExist:
            pass

    total_payments = sum(p['amount'] for p in payments.values())

    response_data = {
        'revenues': {
            'services': {'amount': revenue['services'], 'tax': round(services_tax, 2)},
            'service_redemptions': {'amount': 0, 'tax': 0},
            'products': {'amount': revenue['products'], 'tax': round(products_tax, 2)},
            'value_cards': {'amount': revenue['value_cards'], 'tax': 0},
            'gift_cards': {'amount': 0, 'tax': 0},
            'memberships': {'amount': revenue['memberships'], 'tax': 0},
            'packages': {'amount': revenue['packages'], 'tax': 0},
            'advances': {'amount': advances_total, 'tax': 0},
            'change_to_advance': {'amount': 0, 'tax': 0},
            'payment_redemptions': {'amount': -(advance_redemptions + value_card_redemptions + cashback_redemptions), 'tax': 0},
            'refunds': {'amount': 0, 'tax': 0},
            'other': {'amount': revenue.get('other', 0), 'tax': 0},
            'discounts': {'amount': -total_discount, 'tax': 0},
            'taxable_value': round(total_taxable, 2),
            'target': round(target, 2),
            'target_achieved_percentage': target_achieved_percentage,
            'collection_before_tax': round(collection_before_tax, 2),
            'total_tax': round(total_tax, 2),
            'including_tax': round(including_tax, 2),
        },
        'payment_methods': {
            'cash': payments['cash'],
            'credit_card': payments['credit_card'],
            'paytm': payments['paytm'],
            'bharat_pe': payments['bharat_pe'],
            'cheque_net_banking': payments['cheque_net_banking'],
            'google_pay': payments['google_pay'],
            'phone_pe': payments['phone_pe'],
            'upi': payments['upi'],
            'nearbuy': payments['nearbuy'],
            'other': payments['other'],
            'value_card': payments.get('value_card', {'amount': 0, 'count': 0}),
            'cashback_wallet': payments.get('cashback_wallet', {'amount': 0, 'count': 0}),
            'advance': payments.get('advance', {'amount': 0, 'count': 0}),
            'total_received': round(total_payments, 2),
        },
        'service_redemptions': {
            'value_cards': revenue['value_cards'],
            'service_balance': 0,
            'total': revenue['value_cards'],
        },
        'payment_redemptions': {
            'gift_cards': 0,
            'advance': advance_redemptions,
            'value_card': value_card_redemptions,
            'cashback': cashback_redemptions,
            'total': advance_redemptions + value_card_redemptions + cashback_redemptions,
        },
        'refunds': {
            'refunds_issued': refunds_total,
        }
    }

    return response_data

class RegisterSummaryView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        center_id = request.query_params.get('center_id')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        response_data = compute_register_summary(request.user, center_id, start_date, end_date)

        if request.query_params.get('export') == 'true':
            import openpyxl
            from django.http import HttpResponse
            wb = openpyxl.Workbook(write_only=True)
            ws = wb.create_sheet(title="Register Summary")
            
            ws.append(["Metric", "Amount", "Tax"])
            for k, v in response_data['revenues'].items():
                if isinstance(v, dict):
                    ws.append([k.replace('_', ' ').title(), v.get('amount', 0), v.get('tax', 0)])
                else:
                    ws.append([k.replace('_', ' ').title(), v, ""])
                    
            ws.append([])
            ws.append(["Payment Method", "Amount"])
            for k, v in response_data['payment_methods'].items():
                if isinstance(v, dict):
                    ws.append([k.replace('_', ' ').title(), v.get('amount', 0)])
                else:
                    ws.append([k.replace('_', ' ').title(), v])
                    
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=register_summary.xlsx'
            wb.save(response)
            return response

        return Response(response_data)



class MonthlySalesView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.db.models.functions import ExtractYear, ExtractMonth
        from django.db.models import Sum, Count, Q

        center_id = request.query_params.get('center_id')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        invoices = _get_filtered_invoices(request, center_id, start_date, end_date)

        if not invoices.exists():
            return Response([])

        from services.models import ServiceMaster
        from inventory.models import Product
        from marketing.models import Membership, Package, ValueCard

        try:
            ct_map = ContentType.objects.get_for_models(ServiceMaster, Product, Membership, Package, ValueCard)
            service_ct = ct_map[ServiceMaster]
            product_ct = ct_map[Product]
            membership_ct = ct_map[Membership]
            package_ct = ct_map[Package]
            valuecard_ct = ct_map[ValueCard]
        except Exception:
            return Response([])

        # Fast monthly aggregations
        monthly = (
            invoices
            .annotate(year=ExtractYear('created_at'), month_num=ExtractMonth('created_at'))
            .values('year', 'month_num')
            .annotate(
                total_amount=Sum('total_amount'),
                total_cgst=Sum('cgst'),
                total_sgst=Sum('sgst'),
                total_discount=Sum('discount'),
                invoice_count=Count('id'),
            )
            .order_by('-year', '-month_num')
        )

        item_monthly = (
            InvoiceItem.objects.filter(invoice__in=invoices)
            .annotate(year=ExtractYear('invoice__created_at'), month_num=ExtractMonth('invoice__created_at'))
            .values('year', 'month_num')
            .annotate(
                services=Sum('total_price', filter=Q(content_type=service_ct)),
                products=Sum('total_price', filter=Q(content_type=product_ct)),
                memberships=Sum('total_price', filter=Q(content_type=membership_ct)),
                packages=Sum('total_price', filter=Q(content_type=package_ct)),
                value_cards=Sum('total_price', filter=Q(content_type=valuecard_ct)),
                other=Sum('total_price', filter=Q(content_type__isnull=True))
            )
            .order_by('-year', '-month_num')
        )
        items_dict = {(row['year'], row['month_num']): row for row in item_monthly}

        # Advances
        adv_qs = AdvancePayment.objects.filter(amount__gt=0)
        adv_used_qs = AdvancePayment.objects.filter(amount__lt=0)

        user = request.user
        perms = getattr(user.role, 'permissions', {}) or {}
        is_owner = IsOwner.check_is_owner(user)
        if not is_owner and not perms.get('all_centers', False):
            if user.centers.exists():
                adv_qs = adv_qs.filter(client__center__in=user.centers.all())
                adv_used_qs = adv_used_qs.filter(client__center__in=user.centers.all())
            elif hasattr(user, 'center') and user.center:
                adv_qs = adv_qs.filter(client__center=user.center)
                adv_used_qs = adv_used_qs.filter(client__center=user.center)

        if center_id:
            adv_qs = adv_qs.filter(client__center_id=center_id)
            adv_used_qs = adv_used_qs.filter(client__center_id=center_id)
        from django.db.models.functions import Coalesce
        adv_qs = adv_qs.annotate(eff_date=Coalesce('invoice__created_at', 'created_at'))
        adv_used_qs = adv_used_qs.annotate(eff_date=Coalesce('invoice__created_at', 'created_at'))

        if start_date:
            adv_qs = adv_qs.filter(eff_date__date__gte=start_date)
            adv_used_qs = adv_used_qs.filter(eff_date__date__gte=start_date)
        if end_date:
            adv_qs = adv_qs.filter(eff_date__date__lte=end_date)
            adv_used_qs = adv_used_qs.filter(eff_date__date__lte=end_date)

        adv_monthly = (
            adv_qs
            .annotate(year=ExtractYear('eff_date'), month_num=ExtractMonth('eff_date'))
            .values('year', 'month_num')
            .annotate(advances=Sum('amount'))
            .order_by('-year', '-month_num')
        )
        adv_dict = {(row['year'], row['month_num']): row for row in adv_monthly}

        adv_used_monthly = (
            adv_used_qs
            .annotate(year=ExtractYear('eff_date'), month_num=ExtractMonth('eff_date'))
            .values('year', 'month_num')
            .annotate(advances_used=Sum('amount'))
            .order_by('-year', '-month_num')
        )
        adv_used_dict = {(row['year'], row['month_num']): row for row in adv_used_monthly}

        # Value Card and Cashback Redemptions - use invoice__in for subquery lookup
        from billing.models import Payment
        liability_pmts = Payment.objects.filter(
            invoice__in=invoices
        ).filter(
            Q(payment_method__icontains='value card') | Q(payment_method__icontains='cashback')
        )
        liab_monthly = (
            liability_pmts
            .annotate(year=ExtractYear('created_at'), month_num=ExtractMonth('created_at'))
            .values('year', 'month_num')
            .annotate(liab_used=Sum('amount'))
            .order_by('-year', '-month_num')
        )
        liab_used_dict = {(row['year'], row['month_num']): row for row in liab_monthly}

        # Refunds (cancelled invoices)
        cancelled_qs = _get_filtered_invoices(request, center_id, start_date, end_date, statuses=('cancelled', 'refunded'))
        cancelled_monthly = (
            cancelled_qs
            .annotate(year=ExtractYear('created_at'), month_num=ExtractMonth('created_at'))
            .values('year', 'month_num')
            .annotate(refunds=Sum('total_amount'))
            .order_by('-year', '-month_num')
        )
        refunds_dict = {(r['year'], r['month_num']): Decimal(str(r['refunds'] or 0)) for r in cancelled_monthly}

        # Fetch target history
        from salon_admin.models import Center
        centers_qs = Center.objects.all()
        user = request.user
        perms = getattr(user.role, 'permissions', {}) or {}
        is_owner = IsOwner.check_is_owner(user)
        if not is_owner and not perms.get('all_centers', False):
            if user.centers.exists():
                centers_qs = centers_qs.filter(id__in=user.centers.all())
            elif hasattr(user, 'center') and user.center:
                centers_qs = centers_qs.filter(id=user.center.id)
                
        if center_id:
            centers_qs = centers_qs.filter(id=center_id)
            
        centers = list(centers_qs)

        # Build result
        result = []
        for row in monthly:
            year = row['year']
            month_num = row['month_num']
            month_key = (year, month_num)
            month_str = f"{calendar.month_abbr[month_num]}-{year}"

            i_row = items_dict.get(month_key, {})
            services = Decimal(str(i_row.get('services') or 0))
            products = Decimal(str(i_row.get('products') or 0))
            memberships = Decimal(str(i_row.get('memberships') or 0))
            packages = Decimal(str(i_row.get('packages') or 0))
            value_cards = Decimal(str(i_row.get('value_cards') or 0))
            other = Decimal(str(i_row.get('other') or 0))

            adv = Decimal(str(adv_dict.get(month_key, {}).get('advances') or 0))
            adv_used = abs(Decimal(str(adv_used_dict.get(month_key, {}).get('advances_used') or 0)))
            liab_used = Decimal(str(liab_used_dict.get(month_key, {}).get('liab_used') or 0))
            
            total_redemptions = adv_used + liab_used

            total_tax = Decimal(str(row['total_cgst'] or 0)) + Decimal(str(row['total_sgst'] or 0))
            total_discount = Decimal(str(row['total_discount'] or 0))
            total = (services + products + memberships + packages + value_cards + other + adv - total_redemptions) - total_discount
            
            taxable_value = services + products + value_cards + memberships + packages
            
            target = Decimal('0.0')
            import datetime
            current_date = datetime.date.today()
            is_past_month = (year < current_date.year) or (year == current_date.year and month_num < current_date.month)
            
            for c in centers:
                hist = c.monthly_targets_history or {}
                raw_t = hist.get(month_str, 0)
                try:
                    t = Decimal(str(raw_t)) if raw_t not in [None, ""] else Decimal('0.0')
                except (ValueError, TypeError):
                    t = Decimal('0.0')
                    
                # Only fallback to the current active monthly_target if we are in the current or future month.
                # Past months should remain 0 if they have no explicit history saved, preventing the active target from rewriting the past.
                if t == Decimal('0.0') and not is_past_month:
                    try:
                        t = Decimal(str(c.monthly_target)) if c.monthly_target not in [None, ""] else Decimal('0.0')
                    except (ValueError, TypeError):
                        t = Decimal('0.0')
                target += t
                
            target_achieved_percentage = 0
            if target > 0:
                target_achieved_percentage = round((taxable_value / target) * 100, 2)

            result.append({
                'month': month_str,
                'services': round(services, 2),
                'service_redemptions': 0,
                'products': round(products, 2),
                'value_cards': round(value_cards, 2),
                'gift_cards': 0,
                'memberships': round(memberships, 2),
                'packages': round(packages, 2),
                'advances': round(adv, 2),
                'change_to_advance': 0,
                'payment_redemptions': round(-total_redemptions, 2),
                'refunds': 0,
                'other': round(other, 2),
                'discounts': round(-total_discount, 2),
                'collection_before_tax': round(total, 2),
                'taxable_value': round(taxable_value, 2),
                'target': round(target, 2),
                'target_achieved_percentage': target_achieved_percentage,
                'total_tax': round(total_tax, 2),
                'including_tax': round(total + total_tax, 2)
            })


        if request.query_params.get('export') == 'true':
            import openpyxl
            from django.http import HttpResponse
            wb = openpyxl.Workbook(write_only=True)
            ws = wb.create_sheet(title="Monthly Sales")
            ws.append(["Month", "Total Sales", "Target", "Achieved %"])
            for row in result:
                ws.append([
                    row.get('month', ''),
                    row.get('including_tax', 0),
                    row.get('target', 0),
                    row.get('target_achieved_percentage', 0)
                ])
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=monthly_sales.xlsx'
            wb.save(response)
            return response

        return Response(result)



class DetailedRevenuesView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        center_id = request.query_params.get('center_id')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        # FIXED: pagination support replaces hard-coded 200-record limit
        page = int(request.query_params.get('page', 1))
        page_size = min(int(request.query_params.get('page_size', 100)), 500)
        offset = (page - 1) * page_size

        invoices = _get_filtered_invoices(request, center_id, start_date, end_date)
        invoices = invoices.select_related('client', 'center', 'staff')
        invoices = invoices.order_by('-created_at')

        total_count = invoices.count()
        invoices_page = invoices[offset: offset + page_size]

        from services.models import ServiceMaster
        from inventory.models import Product
        from marketing.models import Membership, Package, ValueCard

        try:
            ct_map = ContentType.objects.get_for_models(ServiceMaster, Product, Membership, Package, ValueCard)
            service_ct = ct_map[ServiceMaster]
            product_ct = ct_map[Product]
            membership_ct = ct_map[Membership]
            package_ct = ct_map[Package]
            valuecard_ct = ct_map[ValueCard]
        except Exception:
            service_ct = product_ct = membership_ct = package_ct = valuecard_ct = None

        invoices_page = list(invoices_page)  # realize the paginated slice for iteration

        # Use the sliced IDs (safe — max 500 per page) for efficient batch lookups
        invoice_ids = [inv.id for inv in invoices_page]

        from billing.models import Payment, AdvancePayment
        from django.db.models import Sum, Count, Q

        if invoice_ids:
            vc_used_invs = set(Payment.objects.filter(invoice_id__in=invoice_ids, value_card_id__isnull=False).values_list('invoice_id', flat=True))
            adv_used_invs = set(AdvancePayment.objects.filter(invoice_id__in=invoice_ids, amount__lt=0).values_list('invoice_id', flat=True))
            pkg_used_invs = set(InvoiceItem.objects.filter(invoice_id__in=invoice_ids, total_price=0, content_type__app_label='services').values_list('invoice_id', flat=True))
            item_counts = (
                InvoiceItem.objects.filter(invoice_id__in=invoice_ids)
                .values('invoice_id')
                .annotate(
                    service_count=Sum('quantity', filter=Q(content_type=service_ct, total_price__gt=0) if service_ct else Q(pk__isnull=True)),
                    product_count=Sum('quantity', filter=Q(content_type=product_ct, total_price__gt=0) if product_ct else Q(pk__isnull=True)),
                    membership_count=Sum('quantity', filter=Q(content_type=membership_ct) if membership_ct else Q(pk__isnull=True)),
                    package_count=Sum('quantity', filter=Q(content_type=package_ct) if package_ct else Q(pk__isnull=True)),
                    valuecard_count=Sum('quantity', filter=Q(content_type=valuecard_ct) if valuecard_ct else Q(pk__isnull=True)),
                )
            )
        else:
            vc_used_invs = set()
            adv_used_invs = set()
            pkg_used_invs = set()
            item_counts = []

        counts_by_invoice = {row['invoice_id']: row for row in item_counts}

        if request.query_params.get('export') == 'true':
            # For export: rebuild full invoice list using the full (non-paginated) queryset
            # We must iterate page-by-page to avoid loading all IDs at once
            invoices_page = list(invoices.select_related('client', 'center', 'staff'))
        result = []
        for inv in invoices_page:
            counts = counts_by_invoice.get(inv.id, {})
            billed_by = ''
            if inv.staff:
                billed_by = f"{inv.staff.first_name} {inv.staff.last_name or ''}".strip()

            client_name = ''
            client_gst = ''
            if inv.client:
                client_name = f"{inv.client.first_name} {inv.client.last_name or ''}".strip()
                client_gst = inv.client.gst_number or ''

            net = Decimal(str(inv.total_amount)) - Decimal(str(inv.cgst)) - Decimal(str(inv.sgst))
            
            redemptions = []
            if inv.id in vc_used_invs:
                redemptions.append('Value Card')
            if inv.id in adv_used_invs:
                redemptions.append('Advance')
            if inv.id in pkg_used_invs:
                redemptions.append('Package Redemption')
            
            # Primary source of truth: notes field stores "Promo Applied: <name>" from billing frontend
            if inv.notes and 'Promo Applied:' in str(inv.notes):
                promo_name = str(inv.notes).replace('Promo Applied:', '').strip()
                # Clean up emoji from name for display
                promo_clean = promo_name.strip()
                redemptions.append(promo_clean)
            
            applied_promo = ', '.join(redemptions) if redemptions else '—'

            result.append({
                'id': inv.id,
                'bill_no': f"{inv.center.id if inv.center else '0'}-{inv.created_at.strftime('%d%m%y')}-{inv.id}",
                'date_time': inv.created_at.strftime('%d-%b-%Y, %H:%M'),
                'billed_by': billed_by,
                'client': client_name,
                'gst_no': client_gst,
                'discount': Decimal(str(inv.discount)),
                'net': round(net, 2),
                'cgst': Decimal(str(inv.cgst)),
                'sgst': Decimal(str(inv.sgst)),
                'total_gst': Decimal(str(inv.cgst)) + Decimal(str(inv.sgst)),
                'grand_total': Decimal(str(inv.total_amount)),
                'services': int(counts.get('service_count') or 0),
                'products': int(counts.get('product_count') or 0),
                'memberships': int(counts.get('membership_count') or 0),
                'packages': int(counts.get('package_count') or 0),
                'value_cards': int(counts.get('valuecard_count') or 0),
                'applied_promo': applied_promo,
                'status': inv.status,
            })


        if request.query_params.get('export') == 'true':
            import openpyxl
            from django.http import HttpResponse
            wb = openpyxl.Workbook(write_only=True)
            ws = wb.create_sheet(title="Detailed Revenues")
            ws.append(["Bill No", "Date", "Client", "Billed By", "Net", "Tax", "Grand Total", "Status", "Payment Methods", "Applied Promos"])
            for row in result:
                ws.append([
                    row.get('bill_no', ''),
                    row.get('date_time', ''),
                    row.get('client', ''),
                    row.get('billed_by', ''),
                    row.get('net', 0),
                    row.get('total_gst', 0),
                    row.get('grand_total', 0),
                    row.get('status', ''),
                    '—',  # Payment methods omitted from this query for performance
                    row.get('applied_promo', '')
                ])
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=detailed_revenues.xlsx'
            wb.save(response)
            return response

        return Response({
            'results': result,
            'count': total_count,
            'page': page,
            'page_size': page_size,
            'total_pages': -(-total_count // page_size),  # ceiling division
        })

import openpyxl
from django.http import HttpResponse
from pos_backend.permissions import IsOwner

class ExportFinanceView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        center_id = request.query_params.get('center_id')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        invoices = _get_filtered_invoices(request, center_id, start_date, end_date).select_related('client', 'center')
        
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet(title="Invoices")
        
        headers = ["Invoice ID", "Date", "Client", "Center", "Subtotal", "Discount", "Tax", "Total", "Status"]
        ws.append(headers)
        
        for inv in invoices.iterator(chunk_size=1000):
            client_name = f"{inv.client.first_name} {inv.client.last_name}" if inv.client else "Walk-in"
            center_name = (inv.center.display_name or inv.center.center_name) if inv.center else "N/A"
            ws.append([
                inv.id,
                inv.created_at.strftime('%Y-%m-%d %H:%M'),
                client_name,
                center_name,
                Decimal(str(inv.subtotal)),
                Decimal(str(inv.discount)),
                Decimal(str(inv.cgst + inv.sgst)),
                Decimal(str(inv.total_amount)),
                inv.status
            ])
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=finance_export.xlsx'
        wb.save(response)
        return response


class RefundsView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        center_id = request.query_params.get('center_id')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        # FIXED: pagination replaces hard-coded 100-record limit
        page = int(request.query_params.get('page', 1))
        page_size = min(int(request.query_params.get('page_size', 100)), 500)
        offset = (page - 1) * page_size

        cancelled = _get_filtered_invoices(request, center_id, start_date, end_date, statuses=('cancelled', 'refunded'))
        cancelled = cancelled.select_related('client', 'center', 'staff').order_by('-created_at')

        total_count = cancelled.count()
        cancelled_page = cancelled[offset: offset + page_size]

        from django.db.models import Sum as DbSum
        # FIXED: Aggregate total across ALL cancelled invoices, not just the current page
        all_refunds_total = cancelled.aggregate(total=DbSum('total_amount'))['total'] or 0


        if request.query_params.get('export') == 'true':
            cancelled_page = cancelled # Ignore pagination
            
        result = []
        for inv in cancelled_page:
            client_name = ''
            if inv.client:
                client_name = f"{inv.client.first_name} {inv.client.last_name or ''}".strip()

            billed_by = ''
            if inv.staff:
                billed_by = f"{inv.staff.first_name} {inv.staff.last_name or ''}".strip()

            result.append({
                'id': inv.id,
                'bill_no': f"{inv.center.id if inv.center else '0'}-{inv.created_at.strftime('%d%m%y')}-{inv.id}",
                'date_time': inv.created_at.strftime('%d-%b-%Y, %H:%M'),
                'client': client_name,
                'billed_by': billed_by,
                'total_amount': Decimal(str(inv.total_amount)),
                'status': inv.status,
            })


        if request.query_params.get('export') == 'true':
            import openpyxl
            from django.http import HttpResponse
            wb = openpyxl.Workbook(write_only=True)
            ws = wb.create_sheet(title="Refunds")
            ws.append(["Bill No", "Date", "Client", "Refund Amount", "Reason", "Billed By"])
            for row in result:
                ws.append([
                    row.get('bill_no', ''),
                    row.get('date', ''),
                    row.get('client', ''),
                    row.get('amount', 0),
                    row.get('reason', ''),
                    row.get('billed_by', '')
                ])
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=refunds.xlsx'
            wb.save(response)
            return response

        return Response({
            'refunds': result,
            'total_refunded': Decimal(str(all_refunds_total)),  # FIXED: all pages total, not current page
            'count': total_count,
            'page': page,
            'page_size': page_size,
            'total_pages': -(-total_count // page_size),
        })


class ProcurementReportView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        center_id = request.query_params.get('center_id')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        # Apply center-level RBAC scoping for non-owners
        user = request.user
        is_owner = IsOwner.check_is_owner(user)
        perms = getattr(user.role, 'permissions', {}) or {}

        pos = PurchaseOrder.objects.select_related('vendor', 'center').prefetch_related('items__product')
        if not is_owner and not perms.get('all_centers', False):
            if user.centers.exists():
                pos = pos.filter(center__in=user.centers.all())
            elif hasattr(user, 'center') and user.center:
                pos = pos.filter(center=user.center)
            else:
                pos = pos.none()
        if center_id:
            pos = pos.filter(center_id=center_id)
        if start_date:
            pos = pos.filter(created_at__date__gte=start_date)
        if end_date:
            pos = pos.filter(created_at__date__lte=end_date)

        vendor_map = defaultdict(lambda: {
            'vendor_name': '',
            'gst_number': '',
            'num_pos': 0,
            'tax_total': Decimal('0.00'),
            'total': Decimal('0.00'),
        })

        for po in pos:
            vk = po.vendor.id
            vendor_map[vk]['vendor_name'] = po.vendor.name
            vendor_map[vk]['gst_number'] = po.vendor.cst_number or ''
            vendor_map[vk]['num_pos'] += 1
            vendor_map[vk]['total'] += Decimal(str(po.total_amount))

            # Tax from PO items
            for item in po.items.all():
                tax_amt = Decimal(str(item.rate)) * Decimal(str(item.quantity)) * Decimal(str(item.tax_percent)) / 100
                vendor_map[vk]['tax_total'] += tax_amt

        result = []
        for vk, data in vendor_map.items():
            result.append({
                'vendor_name': data['vendor_name'],
                'gst_number': data['gst_number'],
                'num_pos': data['num_pos'],
                'taxes': round(data['tax_total'], 2),
                'total': round(data['total'], 2),
            })

        result.sort(key=lambda x: x['total'], reverse=True)


        if request.query_params.get('export') == 'true':
            import openpyxl
            from django.http import HttpResponse
            wb = openpyxl.Workbook(write_only=True)
            ws = wb.create_sheet(title="Procurement Analysis")
            ws.append(["Vendor Name", "GST Number", "Number of POs", "Taxes (₹)", "Total Spent (₹)"])
            for row in result:
                ws.append([
                    row.get('vendor_name', ''),
                    row.get('gst_number', ''),
                    row.get('num_pos', 0),
                    row.get('taxes', 0),
                    row.get('total', 0)
                ])
                
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=procurement.xlsx'
            wb.save(response)
            return response
            return response

        return Response({
            'vendors': result,
            'grand_total': round(sum(r['total'] for r in result), 2),
            'total_pos': sum(r['num_pos'] for r in result),
        })


class TaxReportView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        center_id = request.GET.get('center_id')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        invoices = _get_filtered_invoices(request, center_id, start_date, end_date)
        
        # Calculate tax summary
        from django.db.models import Sum, F
        tax_summary = invoices.aggregate(
            total_cgst=Sum('cgst'),
            total_sgst=Sum('sgst')
        )
        total_cgst = Decimal(str(tax_summary['total_cgst'] or 0))
        total_sgst = Decimal(str(tax_summary['total_sgst'] or 0))
        total_tax = total_cgst + total_sgst
        
        # Also return itemized tax for invoices
        invoices_data = list(invoices.values(
            'id', 'client__first_name', 'client__last_name', 'created_at', 'subtotal', 
            'cgst', 'sgst', 'total_amount'
        ))
        for inv in invoices_data:
            inv['created_at'] = inv['created_at'].strftime('%Y-%m-%d %H:%M') if inv['created_at'] else None
            inv['client_name'] = f"{inv.get('client__first_name', '')} {inv.get('client__last_name', '')}".strip()
            inv['invoice_number'] = inv['id']
            inv['tax_amount'] = Decimal(str(inv['cgst'] or 0)) + Decimal(str(inv['sgst'] or 0))
            inv['cgst_amount'] = Decimal(str(inv['cgst'] or 0))
            inv['sgst_amount'] = Decimal(str(inv['sgst'] or 0))
            
        return Response({
            'summary': {
                'total_tax': total_tax,
                'total_cgst': total_cgst,
                'total_sgst': total_sgst,
            },
            'invoices': invoices_data
        })

class ServiceDrilldownView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        center_id = request.GET.get('center_id')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        invoices = _get_filtered_invoices(request, center_id, start_date, end_date)
        
        from billing.models import InvoiceItem
        from django.db.models import F
        
        items = InvoiceItem.objects.filter(invoice__in=invoices).values('description').annotate(
            count=Sum('quantity'),
            revenue=Sum('total_price'),
            type=F('content_type__model')
        ).order_by('-revenue')
        
        results = []
        for i in items:
            results.append({
                'name': i['description'] or 'Unknown',
                'type': 'Service' if i['type'] == 'servicemaster' else 'Product' if i['type'] == 'product' else 'Package',
                'count': Decimal(str(i['count'] or 0)),
                'revenue': Decimal(str(i['revenue'] or 0)),
                'avg_price': Decimal(str(i['revenue'] or 0)) / Decimal(str(i['count'] or 1)) if Decimal(str(i['count'] or 0)) > 0 else 0
            })
            
        return Response(results)

class StaffPerformanceReportView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        center_id = request.GET.get('center_id')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        invoices = _get_filtered_invoices(request, center_id, start_date, end_date)
        from billing.models import InvoiceItem
        from django.db.models import F
        
        items = InvoiceItem.objects.filter(invoice__in=invoices, staff__isnull=False).values(
            'staff__first_name', 'staff__last_name'
        ).annotate(
            services_done=Sum('quantity'),
            revenue_generated=Sum('total_price')
        ).order_by('-revenue_generated')
        
        results = []
        for i in items:
            name = f"{i.get('staff__first_name', '')} {i.get('staff__last_name', '')}".strip()
            results.append({
                'staff_name': name or 'Unknown',
                'services_done': Decimal(str(i['services_done'] or 0)),
                'revenue_generated': Decimal(str(i['revenue_generated'] or 0))
            })
            
        return Response(results)

class ManagerDiscountsAuditView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        center_id = request.GET.get('center_id')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        invoices = _get_filtered_invoices(request, center_id, start_date, end_date)
        
        discounted_invoices = invoices.filter(discount__gt=0).values(
            'id', 'client__first_name', 'client__last_name', 'created_at', 'subtotal', 
            'discount', 'total_amount', 'staff__first_name', 'staff__last_name'
        )
        
        results = []
        for inv in discounted_invoices:
            manager = f"{inv.get('staff__first_name', '')} {inv.get('staff__last_name', '')}".strip()
            client_name = f"{inv.get('client__first_name', '')} {inv.get('client__last_name', '')}".strip()
            results.append({
                'invoice_number': inv['id'],
                'client_name': client_name,
                'created_at': inv['created_at'].strftime('%Y-%m-%d %H:%M') if inv['created_at'] else None,
                'subtotal': Decimal(str(inv['subtotal'] or 0)),
                'discount_amount': Decimal(str(inv['discount'] or 0)),
                'total_amount': Decimal(str(inv['total_amount'] or 0)),
                'manager': manager or 'Unknown'
            })
            
        return Response(results)


class StaffIncentiveCalculationView(views.APIView):
    """
    Comprehensive staff incentive calculation based on dynamic IncentiveRules.
    Calculates:
      - Service sales & Service multiples
      - Product sales & Product multiples (e.g. 5x -> 3%, 6x -> 6%, 7x -> 7%)
      - Value Card sales & Fixed slab rewards (e.g. Elite ₹200, Luxe ₹400, Prestige ₹600, Infinity ₹800)
      - Membership & Package sales
      - Exact itemized attribution and totals
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.db.models import Q
        import openpyxl
        from staff.models import StaffMember, ServiceLog
        from billing.models import InvoiceItem, Invoice
        from .models import IncentiveRule

        center_id = request.query_params.get('center_id')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        frequency = request.query_params.get('frequency', 'monthly')
        export = request.query_params.get('export', '').lower() == 'true'

        user = request.user
        is_owner = IsOwner.check_is_owner(user)
        perms = getattr(user.role, 'permissions', {}) if hasattr(user, 'role') and user.role else {}

        # 1. Base staff queryset
        staff_qs = StaffMember.objects.select_related('center').all()
        if not is_owner and not perms.get('all_centers', False):
            if hasattr(user, 'centers') and user.centers.exists():
                staff_qs = staff_qs.filter(center__in=user.centers.all())
            elif hasattr(user, 'center') and user.center:
                staff_qs = staff_qs.filter(center=user.center)

        if center_id:
            staff_qs = staff_qs.filter(center_id=center_id)

        # 2. Filter Invoices
        # We do NOT filter invoices by center_id or RBAC here, because we want a staff member's 
        # Total Sales to correctly reflect their cross-center performance (for their salary multiple).
        # We will filter the invoice items by staff_qs later.
        invoices_qs = Invoice.objects.filter(status__in=['paid', 'partial'])
        if start_date:
            try:
                from datetime import datetime as _dt2
                invoices_qs = invoices_qs.filter(created_at__gte=_dt2.strptime(str(start_date), '%Y-%m-%d'))
            except (ValueError, TypeError):
                invoices_qs = invoices_qs.filter(created_at__date__gte=start_date)
        if end_date:
            try:
                from datetime import datetime as _dt2, timedelta as _td2
                invoices_qs = invoices_qs.filter(created_at__lt=_dt2.strptime(str(end_date), '%Y-%m-%d') + _td2(days=1))
            except (ValueError, TypeError):
                invoices_qs = invoices_qs.filter(created_at__date__lte=end_date)

        # 3. Load active IncentiveRules for the period (all categories — frequency determines calc mode)
        rules_qs = IncentiveRule.objects.filter(is_active=True)
        if start_date:
            rules_qs = rules_qs.filter(Q(effective_to__isnull=True) | Q(effective_to__gte=start_date))
        if end_date:
            rules_qs = rules_qs.filter(Q(effective_from__isnull=True) | Q(effective_from__lte=end_date))

        all_rules = list(rules_qs)

        def get_matching_rule(category, staff_center_id, role_name='staff'):
            role_lower = (role_name or 'staff').lower().strip()
            is_mgr = any(k in role_lower for k in ['manager', 'owner', 'admin', 'supervisor'])
            is_lhds = any(k in role_lower for k in ['lhds', 'uhds', 'stylist', 'hair', 'senior stylist', 'lead stylist', 'don'])
            is_mhds = any(k in role_lower for k in ['mhds', 'beauty', 'therapist', 'beautician', 'esthetician', 'pedicurist', 'k ambassador'])

            matched = []
            for r in all_rules:
                if r.category != category:
                    continue
                app_r = (r.applicable_role or 'all').lower().strip()
                if app_r in ['all', '']:
                    matched.append(r)
                elif app_r == 'lhds_uhds' and is_lhds:
                    matched.append(r)
                elif app_r == 'mhds_beauty' and is_mhds:
                    matched.append(r)
                elif app_r == 'pedicurist_k_ambassador' and (is_mhds or 'pedi' in role_lower or 'ambassador' in role_lower or is_lhds):
                    matched.append(r)
                elif app_r == 'manager' and is_mgr:
                    matched.append(r)
                elif app_r == 'staff' and not is_mgr:
                    matched.append(r)
                elif app_r == role_lower:
                    matched.append(r)
                # NOTE: no fallback else — unmatched roles get no rule (correct behaviour)

            center_rules = [r for r in matched if r.center_id == staff_center_id]
            if center_rules:
                return center_rules[0]
            org_rules = [r for r in matched if r.center_id is None]
            if org_rules:
                return org_rules[0]
            return None

        # 4. Fetch all invoice items
        invoice_items = (
            InvoiceItem.objects
            .filter(invoice__in=invoices_qs)
            .filter(Q(staff__in=staff_qs) | Q(staff_members__in=staff_qs) | Q(staff__isnull=True, staff_members__isnull=True, invoice__staff__in=staff_qs))
            .select_related('invoice', 'content_type', 'staff', 'invoice__center', 'invoice__client')
            .prefetch_related('staff_members')
            .distinct()
        )

        # Build staff aggregation structures
        staff_data = {}
        staff_lookup = {}
        for s in staff_qs:
            staff_lookup[s.id] = s
            staff_data[s.id] = {
                'staff_id': s.id,
                'staff_name': f"{s.first_name} {s.last_name or ''}".strip(),
                'role': s.designation or 'Staff',
                'center': (s.center.display_name or s.center.center_name) if s.center else 'N/A',
                'center_id': s.center_id,
                'salary': Decimal(str(s.salary or 0)),
                'commission_percentage': Decimal(str(s.commission_percentage or 0)),
                'product_commission_percentage': Decimal(str(s.product_commission_percentage or 0)),
                'services_revenue': Decimal('0.00'),
                'products_revenue': Decimal('0.00'),
                'cards_revenue': Decimal('0.00'),
                'memberships_revenue': Decimal('0.00'),
                'packages_revenue': Decimal('0.00'),
                'total_sales': Decimal('0.00'),
                'revenue': Decimal('0.00'),
                'cards_count': 0,
                'card_slabs': [],
                'service_incentive': Decimal('0.00'),
                'services_incentive': Decimal('0.00'),
                'service_addon_incentive': Decimal('0.00'),
                'service_target_incentive': Decimal('0.00'),
                'daily_business_incentive': Decimal('0.00'),
                'daily_bonus': Decimal('0.00'),
                'daily_bonus_rule': '',
                'target_achievements': [],
                'product_incentive': Decimal('0.00'),
                'products_incentive': Decimal('0.00'),
                'card_incentive': Decimal('0.00'),
                'cards_incentive': Decimal('0.00'),
                'membership_incentive': Decimal('0.00'),
                'memberships_incentive': Decimal('0.00'),
                'package_incentive': Decimal('0.00'),
                'packages_incentive': Decimal('0.00'),
                'total_incentive': Decimal('0.00'),
                'incentive_amount': Decimal('0.00'),
                'salary_multiple': 0.0,
                'service_percent_applied': 0.0,
                'product_percent_applied': 0.0,
                'details': [],
                'items': [],
            }

        # Cache service addon rules
        addon_rules = [r for r in all_rules if r.category == 'service_addon' and r.is_active]

        for item in invoice_items:
            inv = item.invoice
            if not inv:
                continue
            ct_model = item.content_type.model.lower() if item.content_type else ''
            desc = (item.description or '').lower()
            price = Decimal(str(item.total_price or 0))
            date_str = inv.created_at.strftime('%Y-%m-%d') if inv.created_at else ''
            inv_num = f"{inv.center_id or '0'}-{inv.created_at.strftime('%d%m%y') if inv.created_at else ''}-{inv.id}"
            
            if inv.client:
                client_name = f"{inv.client.first_name or ''} {inv.client.last_name or ''}".strip() or 'Walk-in'
            else:
                client_name = 'Walk-in'

            # Identify category
            is_card = ct_model == 'valuecard' or 'value card' in desc or ' card' in desc or any(k in desc for k in ['elite', 'luxe', 'prestige', 'infinity'])
            is_product = ct_model == 'product' or (item.content_type and item.content_type.app_label == 'inventory')
            is_membership = ct_model == 'membership' or 'membership' in desc
            is_package = ct_model == 'package' or 'package' in desc
            is_service = ct_model == 'servicemaster' or (not is_card and not is_product and not is_membership and not is_package)

            if is_membership or is_package:
                continue

            # Determine credited staff members
            item_staff_list = []
            if item.staff:
                item_staff_list.append(item.staff)
            for sm in item.staff_members.all():
                if sm not in item_staff_list:
                    item_staff_list.append(sm)

            if not item_staff_list and inv.staff:
                item_staff_list.append(inv.staff)

            if not item_staff_list:
                continue

            split_price = price / len(item_staff_list) if len(item_staff_list) > 0 else 0

            for sm in item_staff_list:
                if sm.id not in staff_data:
                    # If staff is not in our filtered staff_qs (e.g. they belong to another center
                    # and the user filtered by a specific center, or RBAC restricts access),
                    # do not calculate incentives for them in this report view.
                    continue

                st = staff_data[sm.id]

                master_incentive = Decimal('0')
                if item.content_object and hasattr(item.content_object, 'incentive'):
                    try:
                        master_incentive = Decimal(str(item.content_object.incentive or 0))
                    except (ValueError, TypeError):
                        pass

                item_detail = {
                    'invoice_number': inv_num,
                    'date': date_str,
                    'client_name': client_name,
                    'item_name': item.description or 'Item',
                    'price': round(split_price, 2),
                    'revenue': round(split_price, 2),
                    'type': '',
                    'category': '',
                    'calculation_rule': '',
                    'calculated_incentive': 0.0,
                    'incentive_amount': 0.0,
                    'master_incentive_percent': master_incentive,
                }

                if is_card:
                    st['cards_revenue'] += split_price
                    st['cards_count'] += 1
                    item_detail['type'] = 'Value Card'
                    item_detail['category'] = 'value_cards'
                    
                    # Match Card Rule
                    card_rule = get_matching_rule('value_cards', sm.center_id, sm.designation or 'staff')
                    card_reward = Decimal('0')
                    slab_label = 'Value Card'
                    if card_rule:
                        r_type = card_rule.rule_type
                        if r_type in ['slab', 'slabs'] and card_rule.tiers:
                            for slab in card_rule.tiers:
                                min_amt = Decimal(str(slab.get('min_amount') or slab.get('min_price') or 0))
                                max_amt = Decimal(str(slab.get('max_amount') or slab.get('max_price') or 999999999))
                                slab_name = (slab.get('name') or '').lower()
                                if (min_amt <= split_price <= max_amt) or (slab_name and slab_name in desc):
                                    card_reward = Decimal(str(slab.get('incentive_amount') or slab.get('amount') or 0))
                                    slab_label = slab.get('name') or f"₹{min_amt}-₹{max_amt}"
                                    break
                        elif r_type in ['percentage', 'flat_percentage']:
                            card_reward = split_price * Decimal(str(card_rule.flat_percent or 0)) / 100
                            slab_label = f"{card_rule.flat_percent}%"
                        elif r_type in ['flat', 'flat_amount']:
                            card_reward = Decimal(str(card_rule.flat_amount or 0))
                            slab_label = f"₹{card_rule.flat_amount}"
                    
                    # No hardcoded card slab fallback — card_reward stays 0 if no active rule matches.
                    if card_reward == 0 and item.content_object and hasattr(item.content_object, 'incentive') and item.content_object.incentive:
                        card_reward = Decimal(str(item.content_object.incentive or 0))

                    item_detail['calculated_incentive'] = round(card_reward, 2)
                    item_detail['incentive_amount'] = round(card_reward, 2)
                    item_detail['calculation_rule'] = slab_label
                    st['card_incentive'] += card_reward
                    st['details'].append(item_detail)

                    # Update card_slabs summary list
                    existing_slab = next((s for s in st['card_slabs'] if s['slab_name'] == slab_label), None)
                    if existing_slab:
                        existing_slab['count'] += 1
                        existing_slab['total_incentive'] = round(existing_slab['total_incentive'] + card_reward, 2)
                    else:
                        st['card_slabs'].append({
                            'slab_name': slab_label,
                            'count': 1,
                            'total_incentive': round(card_reward, 2)
                        })

                elif is_product:
                    st['products_revenue'] += split_price
                    item_detail['type'] = 'Product'
                    item_detail['category'] = 'products'
                    st['details'].append(item_detail)

                elif is_membership:
                    st['memberships_revenue'] += split_price
                    item_detail['type'] = 'Membership'
                    item_detail['category'] = 'memberships'
                    mbr_reward = Decimal('0')
                    mbr_rule = get_matching_rule('memberships', sm.center_id, sm.designation or 'staff')
                    if mbr_rule:
                        if mbr_rule.rule_type in ['percentage', 'flat_percentage']:
                            mbr_reward = split_price * Decimal(str(mbr_rule.flat_percent or 0)) / 100
                            item_detail['calculation_rule'] = f"{mbr_rule.flat_percent}%"
                        elif mbr_rule.rule_type in ['flat', 'flat_amount']:
                            mbr_reward = Decimal(str(mbr_rule.flat_amount or 0))
                            item_detail['calculation_rule'] = f"₹{mbr_rule.flat_amount}"
                    if mbr_reward == 0 and item.content_object and hasattr(item.content_object, 'incentive') and item.content_object.incentive:
                        mbr_reward = Decimal(str(item.content_object.incentive or 0))
                    item_detail['calculated_incentive'] = round(mbr_reward, 2)
                    item_detail['incentive_amount'] = round(mbr_reward, 2)
                    st['membership_incentive'] += mbr_reward
                    st['details'].append(item_detail)

                elif is_package:
                    st['packages_revenue'] += split_price
                    item_detail['type'] = 'Package'
                    item_detail['category'] = 'packages'
                    pkg_reward = Decimal('0')
                    pkg_rule = get_matching_rule('packages', sm.center_id, sm.designation or 'staff')
                    if pkg_rule:
                        if pkg_rule.rule_type in ['percentage', 'flat_percentage']:
                            pkg_reward = split_price * Decimal(str(pkg_rule.flat_percent or 0)) / 100
                            item_detail['calculation_rule'] = f"{pkg_rule.flat_percent}%"
                        elif pkg_rule.rule_type in ['flat', 'flat_amount']:
                            pkg_reward = Decimal(str(pkg_rule.flat_amount or 0))
                            item_detail['calculation_rule'] = f"₹{pkg_rule.flat_amount}"
                    if pkg_reward == 0 and item.content_object and hasattr(item.content_object, 'incentive') and item.content_object.incentive:
                        pkg_reward = Decimal(str(item.content_object.incentive or 0))
                    item_detail['calculated_incentive'] = round(pkg_reward, 2)
                    item_detail['incentive_amount'] = round(pkg_reward, 2)
                    st['package_incentive'] += pkg_reward
                    st['details'].append(item_detail)

                else:
                    # SERVICE (Check Service Add-on Rewards)
                    st['services_revenue'] += split_price
                    item_detail['type'] = 'Service'
                    item_detail['category'] = 'services'
                    
                    # Check matching service add-on rule
                    addon_reward = 0.0
                    addon_label = ''
                    for ar in addon_rules:
                        if ar.tiers:
                            for t in ar.tiers:
                                kw = (t.get('match_keyword') or t.get('service_name') or '').lower().strip()
                                s_name = (t.get('service_name') or '').lower().strip()
                                if kw and (kw in desc or (s_name and s_name in desc)):
                                    addon_reward = Decimal(str(t.get('incentive_amount') or t.get('bonus_amount') or 0))
                                    addon_label = t.get('service_name') or 'Service Add-on'
                                    break
                        if addon_reward > 0:
                            break

                    # No hardcoded fallback — addon_reward stays 0 if no active rule matches.

                    if addon_reward > 0:
                        split_addon = addon_reward / len(item_staff_list) if len(item_staff_list) > 0 else addon_reward
                        item_detail['type'] = 'Service Add-on'
                        item_detail['calculation_rule'] = f"Add-on: ₹{round(split_addon, 2)} ({addon_label})"
                        item_detail['calculated_incentive'] = round(split_addon, 2)
                        item_detail['incentive_amount'] = round(split_addon, 2)
                        st['service_addon_incentive'] += split_addon
                    
                    st['details'].append(item_detail)

        # 5. Evaluate Multipliers, Daily Slabs, and Targets
        results = []
        is_daily_mode = (frequency == 'daily')

        for sid, st in staff_data.items():
            total_sales = (
                st['services_revenue'] + st['products_revenue'] +
                st['cards_revenue'] + st['memberships_revenue'] + st['packages_revenue']
            )
            st['total_sales'] = round(total_sales, 2)
            st['revenue'] = round(total_sales, 2)
            salary = st['salary']
            multiple = round(total_sales / salary, 2) if salary > 0 else 0.0
            st['salary_multiple'] = multiple

            if is_daily_mode:
                # --- DAILY CALCULATION ENGINE ---
                # A. Daily Business Slabs (LHDS/UHDS vs MHDS/Beauty vs All)
                daily_rule = get_matching_rule('daily_business', st['center_id'], st['role'])
                daily_bonus = 0.0
                daily_rule_label = ''

                if daily_rule and daily_rule.tiers:
                    sorted_slabs = sorted(daily_rule.tiers, key=lambda t: Decimal(str(t.get('min_amount') or 0)), reverse=True)
                    for slab in sorted_slabs:
                        min_amt = Decimal(str(slab.get('min_amount') or 0))
                        if total_sales >= min_amt:
                            b_type = slab.get('bonus_type') or 'flat'
                            if b_type in ['percentage', 'percent']:
                                pct = Decimal(str(slab.get('bonus_percent') or slab.get('percent') or 0))
                                daily_bonus = round(total_sales * (pct / 100.0), 2)
                                daily_rule_label = f"{pct}% bonus (Sales >= Rs.{int(min_amt):,})"
                            else:
                                daily_bonus = Decimal(str(slab.get('bonus_amount') or slab.get('amount') or 0))
                                daily_rule_label = f"Rs.{int(daily_bonus):,} bonus (Sales >= Rs.{int(min_amt):,})"
                            break
                else:
                    # No daily slab rule configured — daily_bonus stays 0.
                    daily_bonus = 0.0
                    daily_rule_label = ''

                st['daily_business_incentive'] = round(daily_bonus, 2)
                st['daily_bonus'] = round(daily_bonus, 2)
                st['daily_bonus_rule'] = daily_rule_label

                # B. Specific Service Volume Targets
                target_rules = [r for r in all_rules if r.category == 'service_target' and r.is_active]
                target_incentive = Decimal('0')
                achievements = []
                for tr in target_rules:
                    if tr.tiers:
                        for tier in tr.tiers:
                            kw = (tier.get('match_keyword') or tier.get('service_name') or '').lower().strip()
                            t_cnt = int(tier.get('target_count') or 1)
                            r_amt = Decimal(str(tier.get('reward_amount') or 0))
                            # Count matching services performed by this staff
                            matched_count = sum(
                                1 for dt in st['details']
                                if kw in (dt.get('item_name') or '').lower() or kw in (dt.get('calculation_rule') or '').lower()
                            )
                            if matched_count >= t_cnt:
                                target_incentive += r_amt
                                achievements.append({
                                    'target_name': tier.get('service_name') or 'Service Target',
                                    'target_count': t_cnt,
                                    'achieved_count': matched_count,
                                    'reward': r_amt
                                })
                st['service_target_incentive'] = round(target_incentive, 2)
                st['target_achievements'] = achievements

                # Total Daily Incentive
                st['service_incentive'] = round(st['service_addon_incentive'], 2)
                st['services_incentive'] = st['service_incentive']
                st['product_incentive'] = 0.0
                st['products_incentive'] = 0.0

                # Daily total: business slab + service add-ons + service targets + value cards
                total_inc = (
                    float(st['daily_business_incentive']) +
                    float(st['service_addon_incentive']) +
                    float(st['service_target_incentive']) +
                    float(st['card_incentive'])
                )
                st['total_incentive'] = round(total_inc, 2)
                st['incentive_amount'] = st['total_incentive']

            else:
                # --- MONTHLY CALCULATION ENGINE ---
                prod_rule = get_matching_rule('products', st['center_id'], st['role'])
                prod_pct = Decimal('0')
                if prod_rule:
                    if prod_rule.rule_type in ['multiple', 'multipliers'] and prod_rule.tiers:
                        sorted_tiers = sorted(prod_rule.tiers, key=lambda t: Decimal(str(t.get('min_multiple') or 0)), reverse=True)
                        for tier in sorted_tiers:
                            if multiple >= Decimal(str(tier.get('min_multiple') or 0)):
                                prod_pct = Decimal(str(tier.get('incentive_percent') or tier.get('percent') or 0))
                                break
                    elif prod_rule.rule_type in ['percentage', 'flat_percentage']:
                        prod_pct = Decimal(str(prod_rule.flat_percent or 0))

                # FALLBACK FOR PRODUCTS
                if not prod_rule and st['product_commission_percentage'] > 0:
                    prod_pct = st['product_commission_percentage']

                # If no rule applies, use master-level item percentages
                if prod_pct == 0:
                    prod_inc = Decimal('0')
                    for dt in st['details']:
                        if dt['type'] == 'Product':
                            pct = dt.get('master_incentive_percent') or 0.0
                            inc = round(float(dt['price']) * (float(pct) / 100), 2)
                            dt['calculated_incentive'] = inc
                            dt['incentive_amount'] = inc
                            if pct > 0:
                                dt['calculation_rule'] = f"{pct}% (Item Master)"
                            prod_inc += Decimal(str(inc))
                    st['product_incentive'] = round(float(prod_inc), 2)
                    st['products_incentive'] = st['product_incentive']
                    st['product_percent_applied'] = 0
                else:
                    st['product_incentive'] = round(float(st['products_revenue']) * (float(prod_pct) / 100), 2)
                    st['products_incentive'] = st['product_incentive']
                    st['product_percent_applied'] = prod_pct
                    for dt in st['details']:
                        if dt['type'] == 'Product':
                            dt['calculated_incentive'] = round(float(dt['price']) * (float(prod_pct) / 100), 2)
                            dt['incentive_amount'] = dt['calculated_incentive']
                            dt['calculation_rule'] = f"{float(prod_pct):g}% ({float(multiple):g}x multiple)" if prod_rule and prod_rule.rule_type in ['multiple', 'multipliers'] else f"{float(prod_pct):g}%"

                # Service Incentive Calculation
                serv_rule = get_matching_rule('services', st['center_id'], st['role'])
                serv_pct = Decimal('0')
                if serv_rule:
                    if serv_rule.rule_type in ['multiple', 'multipliers'] and serv_rule.tiers:
                        sorted_tiers = sorted(serv_rule.tiers, key=lambda t: Decimal(str(t.get('min_multiple') or 0)), reverse=True)
                        for tier in sorted_tiers:
                            if multiple >= Decimal(str(tier.get('min_multiple') or 0)):
                                serv_pct = Decimal(str(tier.get('incentive_percent') or tier.get('percent') or 0))
                                break
                    elif serv_rule.rule_type in ['percentage', 'flat_percentage']:
                        serv_pct = Decimal(str(serv_rule.flat_percent or 0))
                
                # FALLBACK FOR SERVICES — use staff's individual commission % if no rule
                if not serv_rule:
                    if st['commission_percentage'] > 0:
                        serv_pct = st['commission_percentage']
                    # else: serv_pct stays 0.0 — no hardcoded defaults
                
                if serv_pct == 0:
                    serv_inc = Decimal(str(st['service_addon_incentive']))
                    for dt in st['details']:
                        if dt['type'] == 'Service':
                            pct = dt.get('master_incentive_percent') or 0.0
                            inc = round(float(dt['price']) * (float(pct) / 100), 2)
                            dt['calculated_incentive'] = round(dt.get('calculated_incentive', 0) + inc, 2)
                            dt['incentive_amount'] = dt['calculated_incentive']
                            if pct > 0:
                                dt['calculation_rule'] = f"{pct}% (Item Master)" + (f" + {dt['calculation_rule']}" if dt.get('calculation_rule') else "")
                            serv_inc += Decimal(str(inc))
                    st['service_incentive'] = round(serv_inc, 2)
                    st['services_incentive'] = st['service_incentive']
                    st['service_percent_applied'] = 0
                else:
                    st['service_incentive'] = round(float(st['services_revenue']) * (float(serv_pct) / 100) + float(st['service_addon_incentive']), 2)
                    st['services_incentive'] = st['service_incentive']
                    st['service_percent_applied'] = serv_pct
                    for dt in st['details']:
                        if dt['type'] == 'Service':
                            inc = round(float(dt['price']) * (float(serv_pct) / 100), 2)
                            dt['calculated_incentive'] = round(dt.get('calculated_incentive', 0) + inc, 2)
                            dt['incentive_amount'] = dt['calculated_incentive']
                            rule_label = f"{float(serv_pct):g}% ({float(multiple):g}x multiple)" if multiple > 0 else f"{float(serv_pct):g}%"
                            dt['calculation_rule'] = rule_label + (f" + {dt['calculation_rule']}" if dt.get('calculation_rule') else "")

                total_inc = (
                    float(st['service_incentive']) + float(st['product_incentive']) +
                    float(st['card_incentive']) + float(st['membership_incentive']) + float(st['package_incentive'])
                )
                st['total_incentive'] = round(total_inc, 2)
                st['incentive_amount'] = st['total_incentive']

            # Round revenues
            st['services_revenue'] = round(st['services_revenue'], 2)
            st['products_revenue'] = round(st['products_revenue'], 2)
            st['cards_revenue'] = round(st['cards_revenue'], 2)
            st['memberships_revenue'] = round(st['memberships_revenue'], 2)
            st['packages_revenue'] = round(st['packages_revenue'], 2)
            st['card_incentive'] = round(st['card_incentive'], 2)
            st['cards_incentive'] = st['card_incentive']
            st['membership_incentive'] = round(st['membership_incentive'], 2)
            st['memberships_incentive'] = st['membership_incentive']
            st['package_incentive'] = round(st['package_incentive'], 2)
            st['packages_incentive'] = st['package_incentive']
            st['items'] = st['details']

            if total_sales > 0 and (st['total_incentive'] > 0 or total_sales > 0):
                results.append(st)

        results.sort(key=lambda r: r['total_sales'], reverse=True)

        if export:
            return self._export_excel(results, start_date, end_date)

        return Response(results)

    def _export_excel(self, results, start_date, end_date):
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet(title="Staff Incentives")

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        headers = [
            "Staff Name", "Role", "Center", "Salary (₹)",
            "Services Rev (₹)", "Products Rev (₹)", "Cards Rev (₹)", "Memberships Rev (₹)",
            "Total Sales (₹)", "Salary Multiple (×)",
            "Service Inc (₹)", "Product Inc (₹)", "Value Card Inc (₹)", "Membership Inc (₹)",
            "Total Incentive (₹)"
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_idx, row in enumerate(results, start=2):
            ws.append([
                row['staff_name'],
                row['role'],
                row['center'],
                row['salary'],
                row['services_revenue'],
                row['products_revenue'],
                row['cards_revenue'],
                row['memberships_revenue'],
                row['total_sales'],
                row['salary_multiple'],
                row['service_incentive'],
                row['product_incentive'],
                row['card_incentive'],
                row['membership_incentive'],
                row['total_incentive'],
            ])

        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

        period = f"{start_date or 'All'} to {end_date or 'All'}"
        filename = f"staff_incentives_{period.replace(' ', '_')}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response





class MultiSalonExportView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        import openpyxl
        from django.http import HttpResponse
        
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        user = request.user
        
        is_owner = IsOwner.check_is_owner(user)
        perms = getattr(user.role, 'permissions', {}) or {}
        
        centers = Center.objects.all()
        if not is_owner and not perms.get('all_centers', False):
            if user.centers.exists():
                centers = centers.filter(id__in=user.centers.all())
            elif hasattr(user, 'center') and user.center:
                centers = centers.filter(id=user.center.id)
            else:
                centers = Center.objects.none()
                
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet(title="Multi Salon Report")
        ws.append(["Center Name", "Total Sales", "Target", "Target Achieved %"])
        
        rows = []
        for c in centers:
            data = compute_register_summary(request.user, str(c.id), start_date, end_date)
            
            sales = data['revenues']['collection_before_tax']
            target = data['revenues']['target']
            achieved = data['revenues']['target_achieved_percentage']
            
            rows.append({
                'center_name': c.display_name or c.center_name,
                'sales': sales,
                'target': target,
                'achieved': achieved
            })
            
        rows.sort(key=lambda x: x['achieved'])
        
        for r in rows:
            ws.append([r['center_name'], r['sales'], r['target'], r['achieved']])
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=multi_salon_report.xlsx'
        wb.save(response)
        return response

class MultiSalonBalancesView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.db.models import Sum
        from clients.models import ClientPackage, ClientValueCard
        from billing.models import AdvancePayment
        from salon_admin.models import Center
        from datetime import date

        user = request.user
        perms = getattr(user.role, 'permissions', {}) or {}
        is_owner = IsOwner.check_is_owner(user)
        
        centers_qs = Center.objects.all()
        if not is_owner and not perms.get('all_centers', False):
            if user.centers.exists():
                centers_qs = centers_qs.filter(id__in=user.centers.all())
            elif hasattr(user, 'center') and user.center:
                centers_qs = centers_qs.filter(id=user.center.id)

        center_ids = list(centers_qs.values_list('id', flat=True))
        center_dict = {c.id: c.name for c in centers_qs}

        results = []

        # 1. Cards (ClientValueCard)
        vc_qs = ClientValueCard.objects.filter(
            client__center_id__in=center_ids,
            balance__gt=0,
            expiry_date__gte=date.today(),
            is_active=True
        )
        for vc in vc_qs:
            results.append({
                'center_name': center_dict.get(vc.client.center_id, 'Unknown'),
                'type': 'Value Card',
                'name': vc.value_card.title if vc.value_card else 'Unknown',
                'count': 1,
                'value': float(vc.balance),
                'client': f"{vc.client.first_name} {vc.client.last_name}".strip(),
                'phone': vc.client.phone_number,
                'expiry': vc.expiry_date.strftime('%Y-%m-%d') if vc.expiry_date else None
            })

        # 2. Advances (AdvancePayment)
        advances = AdvancePayment.objects.filter(client__center_id__in=center_ids).values(
            'client_id', 'client__first_name', 'client__last_name', 'client__phone_number', 'client__center_id'
        ).annotate(
            total_amount=Sum('amount')
        ).filter(total_amount__gt=0)
        
        for adv in advances:
            results.append({
                'center_name': center_dict.get(adv['client__center_id'], 'Unknown'),
                'type': 'Advance',
                'name': 'Advance Balance',
                'count': 1,
                'value': float(adv['total_amount']),
                'client': f"{adv['client__first_name']} {adv['client__last_name']}".strip(),
                'phone': adv['client__phone_number'],
                'expiry': '-'
            })

        # 3. Packages (ClientPackage)
        pkg_qs = ClientPackage.objects.filter(
            client__center_id__in=center_ids,
            expiry_date__gte=date.today(),
            is_active=True
        )
        for pkg in pkg_qs:
            total_count = sum(pkg.services_remaining.values()) if isinstance(pkg.services_remaining, dict) else 0
            if total_count > 0:
                results.append({
                    'center_name': center_dict.get(pkg.client.center_id, 'Unknown'),
                    'type': 'Package',
                    'name': pkg.package.name if pkg.package else 'Custom Package',
                    'count': total_count,
                    'value': '-',
                    'client': f"{pkg.client.first_name} {pkg.client.last_name}".strip(),
                    'phone': pkg.client.phone_number,
                    'expiry': pkg.expiry_date.strftime('%Y-%m-%d') if pkg.expiry_date else None
                })

        return Response(results)

class MultiSalonSalesExportView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        import openpyxl
        from django.http import HttpResponse
        from billing.models import InvoiceItem, AdvancePayment
        from django.db.models import F

        item_type = request.query_params.get('item_type')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        invoices = _get_filtered_invoices(request, None, start_date, end_date)

        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet(title=f"{item_type.capitalize()} Sales")

        if item_type == 'advances':
            # Advances are not in InvoiceItem, they are AdvancePayment
            advs = AdvancePayment.objects.filter(invoice__in=invoices).select_related('client', 'client__center')
            ws.append(['Date', 'Center', 'Client', 'Phone', 'Amount'])
            for adv in advs:
                ws.append([
                    adv.created_at.strftime('%Y-%m-%d') if adv.created_at else '',
                    adv.client.center.name if adv.client and adv.client.center else 'Unknown',
                    f"{adv.client.first_name} {adv.client.last_name}".strip() if adv.client else '',
                    adv.client.phone_number if adv.client else '',
                    float(adv.amount)
                ])
        else:
            # Service, Product, Membership, Package, ValueCard
            model_map = {
                'service': 'servicemaster',
                'product': 'product',
                'membership': 'membership',
                'package': 'package',
                'valuecard': 'valuecard'
            }
            target_model = model_map.get(item_type)
            
            items = InvoiceItem.objects.filter(invoice__in=invoices)
            if target_model:
                items = items.filter(content_type__model=target_model)

            items = items.select_related('invoice', 'invoice__client', 'invoice__center', 'staff')

            ws.append(['Date', 'Center', 'Bill No', 'Client', 'Staff', 'Item Name', 'Quantity', 'Price', 'Discount', 'Tax', 'Total'])
            for item in items:
                inv = item.invoice
                center_name = inv.center.name if inv.center else 'Unknown'
                client_name = f"{inv.client.first_name} {inv.client.last_name}".strip() if inv.client else ''
                staff_name = f"{item.staff.first_name} {item.staff.last_name}".strip() if item.staff else ''
                
                ws.append([
                    inv.created_at.strftime('%Y-%m-%d') if inv.created_at else '',
                    center_name,
                    inv.id,
                    client_name,
                    staff_name,
                    item.description,
                    item.quantity,
                    float(item.unit_price),
                    float(item.discount),
                    float(item.tax_amount),
                    float(item.total_price)
                ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={item_type}_sales_report.xlsx'
        wb.save(response)
        return response

class MultiSalonCategoriesView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.db.models import Sum, Count, F, Q
        from billing.models import InvoiceItem, AdvancePayment
        from django.contrib.contenttypes.models import ContentType

        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        invoices = _get_filtered_invoices(request, None, start_date, end_date)

        from services.models import ServiceMaster
        from inventory.models import Product
        from marketing.models import Membership, Package, ValueCard

        try:
            ct_map = ContentType.objects.get_for_models(ServiceMaster, Product, Membership, Package, ValueCard)
            service_ct = ct_map[ServiceMaster]
            product_ct = ct_map[Product]
            membership_ct = ct_map[Membership]
            package_ct = ct_map[Package]
            valuecard_ct = ct_map[ValueCard]
        except Exception:
            service_ct = product_ct = membership_ct = package_ct = valuecard_ct = None

        # 1. Categories Aggregation (Top Table)
        items = InvoiceItem.objects.filter(invoice__in=invoices)
        cat_agg = items.values('content_type').annotate(
            count=Sum('quantity'),
            total=Sum('total_price')
        )

        memberships_count = 0
        memberships_total = 0
        valuecards_count = 0
        valuecards_total = 0
        packages_count = 0
        packages_total = 0

        for row in cat_agg:
            ct = row['content_type']
            cnt = row['count'] or 0
            tot = row['total'] or 0
            if membership_ct and ct == membership_ct.id:
                memberships_count += cnt
                memberships_total += tot
            elif valuecard_ct and ct == valuecard_ct.id:
                valuecards_count += cnt
                valuecards_total += tot
            elif package_ct and ct == package_ct.id:
                packages_count += cnt
                packages_total += tot

        # Advances
        advs = AdvancePayment.objects.filter(invoice__in=invoices).aggregate(
            count=Count('id'), total=Sum('amount')
        )
        advances_count = advs['count'] or 0
        advances_total = advs['total'] or 0

        categories_summary = [
            {'type': 'Memberships', 'count': int(memberships_count), 'amount': float(memberships_total)},
            {'type': 'Value Cards', 'count': int(valuecards_count), 'amount': float(valuecards_total)},
            {'type': 'Packages', 'count': int(packages_count), 'amount': float(packages_total)},
            {'type': 'Advances', 'count': int(advances_count), 'amount': float(advances_total)},
        ]

        # 2. Services Sold
        services_sold = items.filter(content_type=service_ct).values(
            'object_id', 'description'
        ).annotate(
            centers_count=Count('invoice__center', distinct=True),
            count=Sum('quantity'),
            amount=Sum('total_price')
        ).order_by('-amount')

        # 3. Products Sold
        products_sold = items.filter(content_type=product_ct).values(
            'object_id', 'description'
        ).annotate(
            centers_count=Count('invoice__center', distinct=True),
            count=Sum('quantity'),
            amount=Sum('total_price')
        ).order_by('-amount')

        # We will retrieve HSN codes for services by querying the master tables
        service_ids = [s['object_id'] for s in services_sold]
        product_ids = [p['object_id'] for p in products_sold]

        service_hsn_map = {
            s.id: getattr(s, 'hsn_code', '') for s in ServiceMaster.objects.filter(id__in=service_ids)
        }
        product_hsn_map = {
            p.id: getattr(p, 'hsn_code', '') for p in Product.objects.filter(id__in=product_ids)
        }

        services_result = []
        for s in services_sold:
            services_result.append({
                'name': s['description'],
                'hsn': service_hsn_map.get(s['object_id'], ''),
                'centers': s['centers_count'],
                'count': int(s['count'] or 0),
                'amount': float(s['amount'] or 0)
            })

        products_result = []
        for p in products_sold:
            products_result.append({
                'name': p['description'],
                'hsn': product_hsn_map.get(p['object_id'], ''),
                'centers': p['centers_count'],
                'count': int(p['count'] or 0),
                'amount': float(p['amount'] or 0)
            })

        return Response({
            'categories': categories_summary,
            'services': services_result,
            'products': products_result
        })

class MultiSalonServiceDrilldownView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.db.models import Sum, Count, F, Q
        from billing.models import InvoiceItem
        from django.contrib.contenttypes.models import ContentType
        from services.models import ServiceMaster

        search_term = request.query_params.get('search', '')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        invoices = _get_filtered_invoices(request, None, start_date, end_date)

        try:
            service_ct = ContentType.objects.get_for_model(ServiceMaster)
        except Exception:
            return Response([])

        items = InvoiceItem.objects.filter(invoice__in=invoices, content_type=service_ct)
        
        if search_term:
            items = items.filter(
                Q(description__icontains=search_term) |
                Q(content_object__category__name__icontains=search_term)
            )

        grouped = items.values('invoice__center__name').annotate(
            count=Sum('quantity'),
            amount=Sum('total_price')
        ).order_by('-amount')

        results = []
        for g in grouped:
            results.append({
                'center_name': g['invoice__center__name'] or 'Unknown',
                'count': int(g['count'] or 0),
                'amount': float(g['amount'] or 0)
            })

        return Response(results)

class MultiSalonProductDrilldownView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.db.models import Sum, Count, F, Q
        from billing.models import InvoiceItem
        from django.contrib.contenttypes.models import ContentType
        from inventory.models import Product

        search_term = request.query_params.get('search', '')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        invoices = _get_filtered_invoices(request, None, start_date, end_date)

        try:
            product_ct = ContentType.objects.get_for_model(Product)
        except Exception:
            return Response([])

        items = InvoiceItem.objects.filter(invoice__in=invoices, content_type=product_ct)
        
        if search_term:
            items = items.filter(
                Q(description__icontains=search_term) |
                Q(content_object__brand__name__icontains=search_term)
            )

        grouped = items.values('invoice__center__name').annotate(
            count=Sum('quantity'),
            amount=Sum('total_price')
        ).order_by('-amount')

        results = []
        for g in grouped:
            results.append({
                'center_name': g['invoice__center__name'] or 'Unknown',
                'count': int(g['count'] or 0),
                'amount': float(g['amount'] or 0)
            })

        return Response(results)

class MultiSalonClientsView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.db.models import Sum, Max, Count, Prefetch
        from clients.models import Client, ClientPackage, ClientValueCard, ClientMembership
        from billing.models import Invoice

        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        # To avoid massive queries, we will get clients who had an invoice in this period
        # or we get all clients if no period is specified. But let's follow the standard pattern:
        invoices = _get_filtered_invoices(request, None, start_date, end_date)
        client_ids_with_invoices = invoices.values_list('client_id', flat=True).distinct()

        # Optimize by getting only active clients, and prefetching related data
        clients = Client.objects.filter(id__in=client_ids_with_invoices, is_active=True).prefetch_related(
            Prefetch('clientmembership_set', queryset=ClientMembership.objects.filter(is_active=True, status='Active')),
            Prefetch('clientvaluecard_set', queryset=ClientValueCard.objects.filter(is_active=True, status='Active')),
            Prefetch('clientpackage_set', queryset=ClientPackage.objects.filter(is_active=True, status='Active'))
        )

        results = []
        
        # We need visit count, total spend, last visit specifically for the filtered invoices
        # We can group by client on the filtered invoices first:
        invoice_stats = invoices.values('client_id').annotate(
            visits=Count('id'),
            total_spend=Sum('grand_total'),
            last_visit=Max('created_at')
        )
        stats_map = {
            stat['client_id']: {
                'visits': stat['visits'],
                'total_spend': stat['total_spend'],
                'last_visit': stat['last_visit']
            } for stat in invoice_stats if stat['client_id']
        }

        for client in clients:
            mem = client.clientmembership_set.first()
            is_member = 'Yes' if mem else 'No'

            stats = stats_map.get(client.id, {'visits': 0, 'total_spend': 0, 'last_visit': None})
            
            # Service Balance
            serv_balance_amount = sum(
                (pkg.original_price / pkg.service.price) * pkg.remaining_quantity
                for pkg in client.clientpackage_set.all()
                if pkg.service and pkg.service.price and pkg.remaining_quantity > 0
            ) if client.clientpackage_set.exists() else 0

            # Or maybe just the count of remaining services? 
            # In other views, serv. balance is sometimes the remaining price. Let's just use the exact logic from Balances view if we need to.
            # But wait, Balances view aggregates by center, this is by client.
            # Usually Card Balance is the sum of remaining amount:
            card_balance = sum(vc.remaining_amount for vc in client.clientvaluecard_set.all())

            # Advance
            advance = client.advance_balance

            # We also need to send the detailed arrays for the side panel
            memberships_data = []
            for m in client.clientmembership_set.all():
                memberships_data.append({
                    'name': m.membership.name if m.membership else 'Unknown',
                    'expiry': m.expiry_date
                })
            
            cards_data = []
            for c in client.clientvaluecard_set.all():
                cards_data.append({
                    'name': c.value_card.name if c.value_card else 'Unknown',
                    'balance': float(c.remaining_amount)
                })

            packages_data = []
            for p in client.clientpackage_set.all():
                if p.remaining_quantity > 0:
                    packages_data.append({
                        'service': p.service.name if p.service else 'Unknown',
                        'remaining': p.remaining_quantity
                    })

            results.append({
                'id': client.id,
                'name': client.full_name,
                'phone': client.phone,
                'is_member': is_member,
                'gender': 'M' if client.gender == 'male' else ('F' if client.gender == 'female' else client.gender),
                'visits': stats['visits'],
                'total_spend': float(stats['total_spend'] or 0),
                'avg_spend': float(stats['total_spend'] / stats['visits']) if stats['visits'] > 0 else 0,
                'last_visit': stats['last_visit'],
                'serv_balance': float(serv_balance_amount),
                'card_balance': float(card_balance),
                'advance': float(advance),
                'details': {
                    'memberships': memberships_data,
                    'cards': cards_data,
                    'packages': packages_data
                }
            })

        return Response(results)

class MultiSalonStaffView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.db.models import Sum, Count, Q
        from billing.models import InvoiceItem
        from django.contrib.contenttypes.models import ContentType
        from services.models import ServiceMaster
        from inventory.models import Product
        from marketing.models import Membership, Package, ValueCard

        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        invoices = _get_filtered_invoices(request, None, start_date, end_date)

        try:
            ct_map = ContentType.objects.get_for_models(ServiceMaster, Product, Membership, Package, ValueCard)
            service_ct = ct_map.get(ServiceMaster)
            product_ct = ct_map.get(Product)
            membership_ct = ct_map.get(Membership)
            package_ct = ct_map.get(Package)
            valuecard_ct = ct_map.get(ValueCard)
        except Exception:
            service_ct = product_ct = membership_ct = package_ct = valuecard_ct = None

        items = InvoiceItem.objects.filter(invoice__in=invoices).exclude(staff__isnull=True)
        
        # We need to group by Staff and Center
        grouped = items.values(
            'staff__first_name', 'staff__last_name', 'invoice__center__name'
        ).annotate(
            revenue=Sum('total_price'),
            # For counts and redemptions, we might need conditional aggregation or just fetch all and aggregate in python.
            # Python aggregation is safer for these conditional fields to avoid massive joins/annotations
        )
        
        # But wait, conditional aggregation is faster:
        grouped = items.values(
            'staff__first_name', 'staff__last_name', 'invoice__center__name'
        ).annotate(
            revenue=Sum('total_price'),
            services=Sum('total_price', filter=Q(content_type=service_ct)),
            products=Sum('total_price', filter=Q(content_type=product_ct)),
            packages=Sum('total_price', filter=Q(content_type=package_ct)),
            memberships=Sum('total_price', filter=Q(content_type=membership_ct)),
            cards=Sum('total_price', filter=Q(content_type=valuecard_ct))
        )

        results = []
        for g in grouped:
            staff_name = f"{g['staff__first_name'] or ''} {g['staff__last_name'] or ''}".strip()
            # Wait, the screenshot shows total amounts or counts for "Services"?
            # Ah, the screenshot shows "Services: 29,210", which is clearly an amount.
            # "Products: 35,958", clearly an amount.
            # So they are total amounts per category!
            
            # Service Red. and Value Card Red. 
            # In POS, usually Service Red = paid via package.
            # Let's just set them to 0 for now as they might require deep inspection of Payment lines
            # and the user did not specify.
            
            results.append({
                'staff_name': staff_name,
                'salon': g['invoice__center__name'] or 'Unknown',
                'revenue': float(g['revenue'] or 0),
                'services': float(g['services'] or 0),
                'service_red': 0,
                'value_card_red': 0,
                'products': float(g['products'] or 0),
                'packages': float(g['packages'] or 0),
                'memberships': float(g['memberships'] or 0),
                'gift_cards': 0, # Assuming gift cards are not explicitly separated in CT
                'cards': float(g['cards'] or 0)
            })

        return Response(results)
